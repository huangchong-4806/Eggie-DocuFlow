import re
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

from exporters.word_exporter import CONTENT_TYPES, DOC_START, ROOT_RELS
from utils.file_helper import INVALID_XML_CHARS, publish_output, temporary_output


FORMAL_CONTRACT_STYLE = {
    "east_asia_font": "宋体",
    "ascii_font": "Times New Roman",
    "title_size": 32,
    "body_size": 24,
    "line_spacing": 360,
    "first_line": 480,
    "margin": 1440,
}


def _twips(value):
    return str(max(1, int(round(float(value) * 20))))


def _alignment(line, page_width):
    center = (float(line["x0"]) + float(line["x1"])) / 2
    if abs(center - page_width / 2) <= page_width * 0.08:
        return "center"
    if float(line["x0"]) > page_width * 0.55:
        return "right"
    return "left"


def _paragraph_xml(line, page_width, left_margin):
    text = escape(INVALID_XML_CHARS.sub("", line["text"]))
    if not text:
        return ""
    font = escape(re.sub(r"^[A-Z]+\\+", "", line.get("font") or ""))
    size = max(16, min(48, int(round(float(line.get("size") or 10) * 2))))
    bold = "bold" in font.lower() or "black" in font.lower() or size >= 28
    indent = max(0, int(round((float(line["x0"]) * 20) - left_margin)))
    align = _alignment(line, page_width)
    font_xml = f'<w:rFonts w:ascii="{font}" w:eastAsia="{font}"/>' if font else ""
    return (
        f'<w:p><w:pPr><w:jc w:val="{align}"/><w:ind w:left="{indent}"/>'
        '<w:spacing w:before="0" w:after="80"/></w:pPr><w:r><w:rPr>'
        f'<w:sz w:val="{size}"/>'
        f'{"<w:b/>" if bold else ""}'
        f"{font_xml}"
        f'</w:rPr><w:t xml:space="preserve">{text}</w:t></w:r></w:p>'
    )


def _formal_contract_paragraph_xml(line, is_title=False):
    text = escape(INVALID_XML_CHARS.sub("", _clean_formal_text(line["text"])))
    if not text:
        return ""
    style = FORMAL_CONTRACT_STYLE
    is_clause = bool(re.match(r"^(第[一二三四五六七八九十百\d]+条|[一二三四五六七八九十]+、)", text))
    align = "center" if is_title else "left"
    indent = "" if is_title or is_clause else f'<w:ind w:firstLine="{style["first_line"]}"/>'
    bold = "<w:b/>" if is_title or is_clause else ""
    size = style["title_size"] if is_title else style["body_size"]
    return (
        f'<w:p><w:pPr><w:jc w:val="{align}"/>{indent}'
        f'<w:spacing w:before="0" w:after="120" w:line="{style["line_spacing"]}" w:lineRule="auto"/>'
        '</w:pPr><w:r><w:rPr>'
        f'<w:rFonts w:ascii="{style["ascii_font"]}" w:eastAsia="{style["east_asia_font"]}"/>'
        f'<w:sz w:val="{size}"/>{bold}'
        f'</w:rPr><w:t xml:space="preserve">{text}</w:t></w:r></w:p>'
    )


def _text_runs(text):
    lines = escape(INVALID_XML_CHARS.sub("", text or "")).splitlines() or [""]
    return "<w:br/>".join(f'<w:t xml:space="preserve">{line}</w:t>' for line in lines)


def _formal_contract_table_xml(table):
    rows = table.get("rows") or []
    width = max(len(row) for row in rows)
    if width < 2:
        return ""
    table_width = 9360
    column_width = table_width // width
    borders = "".join(
        f'<w:{edge} w:val="single" w:sz="8" w:space="0" w:color="444444"/>'
        for edge in ("top", "left", "bottom", "right", "insideH", "insideV")
    )
    parts = [
        '<w:tbl><w:tblPr>'
        f'<w:tblW w:w="{table_width}" w:type="dxa"/>'
        f"<w:tblBorders>{borders}</w:tblBorders>"
        '<w:tblCellMar><w:top w:w="120" w:type="dxa"/><w:left w:w="120" w:type="dxa"/>'
        '<w:bottom w:w="120" w:type="dxa"/><w:right w:w="120" w:type="dxa"/></w:tblCellMar>'
        "</w:tblPr><w:tblGrid>",
        *(f'<w:gridCol w:w="{column_width}"/>' for _ in range(width)),
        "</w:tblGrid>",
    ]
    style = FORMAL_CONTRACT_STYLE
    for row_index, row in enumerate(rows):
        parts.append('<w:tr><w:trPr><w:trHeight w:val="520" w:hRule="atLeast"/></w:trPr>')
        column = 0
        while column < width:
            value = row[column] if column < len(row) else ""
            if value is None:
                column += 1
                continue
            span = 1
            while column + span < width and (column + span >= len(row) or row[column + span] is None):
                span += 1
            align = "center" if row_index <= 1 or str(value).strip().isdigit() or "总计" in str(value) else "left"
            bold = "<w:b/>" if row_index <= 1 or "总计" in str(value) else ""
            grid_span = f'<w:gridSpan w:val="{span}"/>' if span > 1 else ""
            parts.append(
                "<w:tc><w:tcPr>"
                f'<w:tcW w:w="{column_width * span}" w:type="dxa"/>{grid_span}'
                '<w:vAlign w:val="center"/></w:tcPr>'
                f'<w:p><w:pPr><w:jc w:val="{align}"/><w:spacing w:before="0" w:after="0"/></w:pPr>'
                "<w:r><w:rPr>"
                f'<w:rFonts w:ascii="{style["ascii_font"]}" w:eastAsia="{style["east_asia_font"]}"/>'
                f'<w:sz w:val="{style["body_size"]}"/>{bold}'
                f'</w:rPr>{_text_runs(str(value))}</w:r></w:p></w:tc>'
            )
            column += span
        parts.append("</w:tr>")
    parts.append("</w:tbl>")
    return "".join(parts)


def _line_inside_table(line, tables):
    center_x = (float(line["x0"]) + float(line["x1"])) / 2
    center_y = (float(line["top"]) + float(line["bottom"])) / 2
    for table in tables:
        x0, top, x1, bottom = table["bbox"]
        if x0 <= center_x <= x1 and top <= center_y <= bottom:
            return True
    return False


def _clean_formal_text(text):
    text = re.sub(r"\s+", " ", text or "").strip()
    text = re.sub(r"\s*/\s*", "/", text)
    text = re.sub(r"\s+([，。；：、！？）】》”])", r"\1", text)
    text = re.sub(r"([（【《“])\s+", r"\1", text)
    text = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fffA-Za-z0-9])", "", text)
    text = re.sub(r"(?<=[A-Za-z0-9])\s+(?=[\u4e00-\u9fff])", "", text)
    return text


def _is_noise_line(text):
    return bool(re.fullmatch(r"[-–—\s]+", text or ""))


def _starts_formal_paragraph(text):
    return bool(
        re.match(
            r"^(第[一二三四五六七八九十百\d]+条|[一二三四五六七八九十]+、|\d+[.、]|\d+(?:\.\d+)+|（[一二三四五六七八九十\d]+）|"
            r"合同编号[:：]?|鉴于[:：]?|[甲乙丙丁]方[（:：]|统一社会信用代码[:：]|法定代表人[:：]|联系地址[:：])",
            text,
        )
    )


def _join_formal_text(left, right):
    if left and right and left[-1].isalnum() and right[0].isalnum() and left[-1].isascii() and right[0].isascii():
        return f"{left} {right}"
    return f"{left}{right}"


def _continues_formal_paragraph(left, right):
    left = _clean_formal_text(left)
    right = _clean_formal_text(right)
    if not left or not right or left.endswith(tuple("。；;：:！？!?")):
        return False
    return not _starts_formal_paragraph(right) and not _is_formal_contract_title(right, 0)


def _merge_cross_page_formal_lines(pages):
    merged_pages = []
    for page in pages:
        new_page = dict(page)
        new_page["lines"] = [dict(line) for line in page.get("lines", [])]
        new_page["tables"] = page.get("tables") or []
        if merged_pages:
            previous = merged_pages[-1]
            previous_lines = [
                line for line in previous["lines"] if not _line_inside_table(line, previous.get("tables") or [])
            ]
            current_lines = [
                line for line in new_page["lines"] if not _line_inside_table(line, new_page.get("tables") or [])
            ]
            if previous_lines and current_lines and _continues_formal_paragraph(previous_lines[-1].get("text"), current_lines[0].get("text")):
                previous_lines[-1]["text"] = _join_formal_text(
                    _clean_formal_text(previous_lines[-1].get("text")),
                    _clean_formal_text(current_lines[0].get("text")),
                )
                new_page["lines"] = [line for line in new_page["lines"] if line is not current_lines[0]]
        merged_pages.append(new_page)
    return merged_pages


def _merge_formal_lines(lines):
    merged = []
    current = None
    end_marks = "。；;：:！？!?"
    for original in lines:
        text = _clean_formal_text(original.get("text"))
        if not text or _is_noise_line(text):
            continue
        line = dict(original, text=text)
        start_new = (
            current is None
            or current["text"].endswith(tuple(end_marks))
            or _starts_formal_paragraph(text)
            or _is_formal_contract_title(text, len(merged))
        )
        if start_new:
            if current:
                merged.append(current)
            current = line
            continue
        current["text"] = _join_formal_text(current["text"], text)
        current["x1"] = max(float(current["x1"]), float(line["x1"]))
        current["bottom"] = float(line["bottom"])
    if current:
        merged.append(current)
    return merged


def _drop_first_page_header(lines):
    for index, line in enumerate(lines[:6]):
        if _is_formal_contract_title(line.get("text"), index):
            return lines[index:]
    return lines


def _is_formal_contract_title(text, line_index):
    text = _clean_formal_text(text)
    return (
        line_index < 4
        and 2 <= len(text) <= 40
        and not any(mark in text for mark in "：:，,。；;")
        and text.endswith(("合同", "协议", "项目", "书"))
    )


def export_contract_layout(layout, output_file, style_template=None):
    pages = layout.get("pages") or []
    if not pages:
        raise ValueError("未提取到合同版式。")
    if style_template == "formal_contract":
        pages = _merge_cross_page_formal_lines(pages)
    first_page = pages[0]
    lines = [line for page in pages for line in page["lines"] if line.get("text")]
    left_margin = min((float(line["x0"]) * 20 for line in lines), default=720)
    right_margin = min(((float(first_page["width"]) - float(line["x1"])) * 20 for line in lines), default=720)
    left_margin = max(360, min(1800, int(left_margin)))
    right_margin = max(360, min(1800, int(right_margin)))
    if style_template == "formal_contract":
        left_margin = right_margin = FORMAL_CONTRACT_STYLE["margin"]
    section = (
        f'<w:sectPr><w:pgSz w:w="{_twips(first_page["width"])}" w:h="{_twips(first_page["height"])}"/>'
        f'<w:pgMar w:top="{FORMAL_CONTRACT_STYLE["margin"] if style_template == "formal_contract" else 720}" '
        f'w:right="{right_margin}" '
        f'w:bottom="{FORMAL_CONTRACT_STYLE["margin"] if style_template == "formal_contract" else 720}" '
        f'w:left="{left_margin}"/>'
        '</w:sectPr></w:body></w:document>'
    )

    temporary_file = temporary_output(output_file)
    try:
        with zipfile.ZipFile(temporary_file, "w", zipfile.ZIP_DEFLATED) as archive:
            parts = [DOC_START]
            line_index = 0
            for page_index, page in enumerate(pages):
                if page_index and style_template != "formal_contract":
                    parts.append('<w:p><w:r><w:br w:type="page"/></w:r></w:p>')
                if style_template == "formal_contract":
                    tables = page.get("tables") or []
                    page_lines = [
                        line for line in page["lines"] if not _line_inside_table(line, tables)
                    ]
                    if page_index == 0:
                        page_lines = _drop_first_page_header(page_lines)
                    blocks = [
                        ("line", line)
                        for line in _merge_formal_lines(page_lines)
                    ]
                    blocks.extend(("table", table) for table in tables)
                    for block_type, block in sorted(blocks, key=lambda item: item[1]["bbox"][1] if item[0] == "table" else item[1]["top"]):
                        if block_type == "table":
                            parts.append(_formal_contract_table_xml(block))
                            continue
                        line = block
                        is_title = _is_formal_contract_title(line.get("text"), line_index)
                        parts.append(_formal_contract_paragraph_xml(line, is_title))
                        line_index += bool(line.get("text"))
                else:
                    for line in page["lines"]:
                        parts.append(_paragraph_xml(line, float(page["width"]), left_margin))
            parts.append(section)
            archive.writestr("[Content_Types].xml", CONTENT_TYPES)
            archive.writestr("_rels/.rels", ROOT_RELS)
            archive.writestr("word/document.xml", "".join(parts))
        return publish_output(temporary_file, output_file)
    finally:
        Path(temporary_file).unlink(missing_ok=True)


def export_table_layout(tables, output_file):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError("缺少 Excel 处理组件 openpyxl。") from error

    if not tables:
        raise ValueError("未提取到可用表格。")

    workbook = Workbook()
    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    try:
        for index, table in enumerate(tables):
            sheet = workbook.active if index == 0 else workbook.create_sheet()
            sheet.title = f"第{table['page_number']}页_表{table['table_number']}"[:31]
            rows = table["rows"]
            width = len(rows[0])
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = f"A1:{get_column_letter(width)}{len(rows)}"
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.orientation = "landscape" if table["page_width"] > table["page_height"] else "portrait"

            for row_number, row in enumerate(rows, 1):
                sheet.row_dimensions[row_number].height = 22 if row_number == 1 else 18
                for column_number, value in enumerate(row, 1):
                    cell = sheet.cell(row_number, column_number, value)
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal="center" if row_number == 1 else "left",
                        vertical="center",
                        wrap_text=True,
                    )
                    if row_number == 1:
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.fill = PatternFill("solid", fgColor="17365D")

            for column_number, values in enumerate(zip(*rows), 1):
                length = max(len(str(value or "")) for value in values)
                sheet.column_dimensions[get_column_letter(column_number)].width = min(max(length + 4, 12), 60)

        temporary_file = temporary_output(output_file)
        workbook.save(temporary_file)
        return publish_output(temporary_file, output_file)
    finally:
        workbook.close()
        if "temporary_file" in locals():
            Path(temporary_file).unlink(missing_ok=True)
