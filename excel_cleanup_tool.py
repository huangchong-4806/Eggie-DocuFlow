import re
from bisect import bisect_left
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

from utils.file_helper import available_output_path, publish_output, temporary_output


SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}


@dataclass(frozen=True)
class CleanupOptions:
    sheet_name: str
    header_row: int = 1
    remove_empty_rows: bool = True
    trim_whitespace: bool = True
    deduplicate: bool = False
    duplicate_columns: tuple = ()
    normalize_dates: bool = False
    normalize_numbers: bool = False


@dataclass(frozen=True)
class CleanupPreview:
    source_file: str
    sheet_name: str
    headers: tuple
    original_rows: int
    original_columns: int
    blank_rows: int
    duplicate_rows: int
    whitespace_cells: int
    date_cells: int
    number_cells: int
    formula_cells: int


@dataclass(frozen=True)
class CleanupResult:
    output_file: str
    log_file: str
    preview: CleanupPreview
    final_rows: int
    removed_rows: int


def _source_path(source_file):
    source = Path(source_file).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Excel 文件不存在：{source}")
    if source.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise ValueError("只支持 .xlsx 和 .xlsm 文件。")
    return source


def workbook_sheet_names(source_file):
    source = _source_path(source_file)
    workbook = load_workbook(
        source,
        read_only=True,
        data_only=False,
        keep_vba=source.suffix.lower() == ".xlsm",
    )
    try:
        return tuple(workbook.sheetnames)
    finally:
        workbook.close()


def _clean_text(value):
    if not isinstance(value, str) or value.startswith("="):
        return value
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in value.splitlines()]
    return "\n".join(lines).strip()


def _is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _row_values(row, trim_whitespace):
    values = []
    for cell in row:
        value = cell.value
        values.append(_clean_text(value) if trim_whitespace else value)
    return tuple(values)


def _selected_key(values, columns):
    if not columns:
        return values
    return tuple(values[index - 1] if index <= len(values) else None for index in columns)


def _sheet_preview(source, sheet, options):
    if sheet.max_row is None or sheet.max_column is None:
        sheet.calculate_dimension(force=True)
    max_row = max(sheet.max_row or 1, 1)
    max_column = max(sheet.max_column or 1, 1)
    if options.header_row < 1 or options.header_row > max_row:
        raise ValueError("表头行超出工作表范围。")
    header_cells = next(
        sheet.iter_rows(
            min_row=options.header_row,
            max_row=options.header_row,
            max_col=max_column,
        )
    )
    headers = tuple(
        str(cell.value).strip() if cell.value is not None else f"第 {index} 列"
        for index, cell in enumerate(header_cells, 1)
    )

    blank_rows = 0
    duplicate_rows = 0
    whitespace_cells = 0
    date_cells = 0
    number_cells = 0
    formula_cells = 0
    seen = set()
    for row_number, row in enumerate(sheet.iter_rows(), 1):
        for cell in row:
            value = cell.value
            if isinstance(value, str):
                if value.startswith("="):
                    formula_cells += 1
                elif _clean_text(value) != value:
                    whitespace_cells += 1
            elif isinstance(value, (datetime, date)):
                date_cells += 1
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                number_cells += 1

        if row_number <= options.header_row:
            continue
        values = _row_values(row, options.trim_whitespace)
        if all(_is_blank(value) for value in values):
            blank_rows += 1
            continue
        if options.deduplicate:
            key = _selected_key(values, options.duplicate_columns)
            if key in seen:
                duplicate_rows += 1
            else:
                seen.add(key)

    return CleanupPreview(
        str(source),
        options.sheet_name,
        headers,
        max(0, max_row - options.header_row),
        max_column,
        blank_rows,
        duplicate_rows,
        whitespace_cells,
        date_cells,
        number_cells,
        formula_cells,
    )


def preview_cleanup(source_file, options):
    source = _source_path(source_file)
    workbook = load_workbook(
        source,
        read_only=True,
        data_only=False,
        keep_vba=source.suffix.lower() == ".xlsm",
    )
    try:
        if options.sheet_name not in workbook.sheetnames:
            raise ValueError("选择的工作表不存在。")
        return _sheet_preview(source, workbook[options.sheet_name], options)
    finally:
        workbook.close()


def _apply_cleanup(sheet, options, progress_callback=None):
    total_rows = max(sheet.max_row, 1)
    if options.trim_whitespace or options.normalize_dates or options.normalize_numbers:
        for row_number, row in enumerate(sheet.iter_rows(), 1):
            for cell in row:
                value = cell.value
                if options.trim_whitespace and isinstance(value, str) and not value.startswith("="):
                    cell.value = _clean_text(value)
                    value = cell.value
                if options.normalize_dates and isinstance(value, datetime):
                    cell.number_format = (
                        "yyyy-mm-dd hh:mm:ss"
                        if value.time() != time(0, 0)
                        else "yyyy-mm-dd"
                    )
                elif options.normalize_dates and isinstance(value, date):
                    cell.number_format = "yyyy-mm-dd"
                if options.normalize_numbers and isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.number_format = "#,##0" if float(value).is_integer() else "#,##0.00"
            if progress_callback and row_number % 200 == 0:
                progress_callback(row_number, total_rows, f"正在清理第 {row_number} 行")

    formulas = {
        cell.coordinate: cell.value
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }
    remove_rows = []
    seen = set()
    for row in sheet.iter_rows(
        min_row=options.header_row + 1,
        max_row=sheet.max_row,
        max_col=sheet.max_column,
    ):
        values = tuple(cell.value for cell in row)
        if all(_is_blank(value) for value in values):
            if options.remove_empty_rows:
                remove_rows.append(row[0].row)
            continue
        if options.deduplicate:
            key = _selected_key(values, options.duplicate_columns)
            if key in seen:
                remove_rows.append(row[0].row)
            else:
                seen.add(key)

    remove_rows.sort()
    for row_number in reversed(remove_rows):
        sheet.delete_rows(row_number, 1)
    removed = set(remove_rows)
    for coordinate, formula in formulas.items():
        original_row, original_column = coordinate_to_tuple(coordinate)
        if original_row in removed:
            continue
        new_row = original_row - bisect_left(remove_rows, original_row)
        if new_row != original_row:
            destination = f"{get_column_letter(original_column)}{new_row}"
            sheet[destination] = Translator(
                formula,
                origin=coordinate,
            ).translate_formula(destination)
    if progress_callback:
        progress_callback(total_rows, total_rows, "数据清理完成，正在保存新文件")
    return len(remove_rows)


def _write_log(log_file, source, options, preview, output_file, final_rows, removed_rows):
    with Path(log_file).open("w", encoding="utf-8") as handle:
        handle.write("Eggie DocuFlow Excel 数据清理日志\n")
        handle.write(f"来源文件：{source}\n")
        handle.write(f"工作表：{options.sheet_name}\n")
        handle.write(f"表头行：{options.header_row}\n")
        handle.write(f"删除空白行：{options.remove_empty_rows}\n")
        handle.write(f"清理文字空格：{options.trim_whitespace}\n")
        handle.write(f"排重：{options.deduplicate}\n")
        handle.write(f"排重列：{','.join(map(str, options.duplicate_columns)) or '整行'}\n")
        handle.write(f"统一日期显示：{options.normalize_dates}\n")
        handle.write(f"统一数字显示：{options.normalize_numbers}\n\n")
        handle.write("匹配结果：\n")
        handle.write(f"original_rows={preview.original_rows}\n")
        handle.write(f"blank_rows={preview.blank_rows}\n")
        handle.write(f"duplicate_rows={preview.duplicate_rows}\n")
        handle.write(f"whitespace_cells={preview.whitespace_cells}\n")
        handle.write(f"date_cells={preview.date_cells}\n")
        handle.write(f"number_cells={preview.number_cells}\n")
        handle.write(f"formula_cells={preview.formula_cells}\n\n")
        handle.write("计算过程：\n")
        handle.write(f"removed_rows={removed_rows}\n")
        handle.write(f"final_rows={final_rows}\n\n")
        handle.write("文件生成状态：\n")
        handle.write(f"output_file={output_file}\n")
        handle.write(f"log_file={log_file}\n")


def clean_workbook(
    source_file,
    output_folder,
    options,
    progress_callback=None,
):
    source = _source_path(source_file)
    output_folder = Path(output_folder).expanduser().resolve()
    output_folder.mkdir(parents=True, exist_ok=True)
    preview = preview_cleanup(source, options)
    keep_vba = source.suffix.lower() == ".xlsm"
    workbook = load_workbook(source, data_only=False, keep_vba=keep_vba)
    temporary_file = None
    try:
        if options.sheet_name not in workbook.sheetnames:
            raise ValueError("选择的工作表不存在。")
        sheet = workbook[options.sheet_name]
        removed_rows = _apply_cleanup(sheet, options, progress_callback)
        final_rows = max(0, sheet.max_row - options.header_row)
        output_file = available_output_path(
            output_folder / f"{source.stem}_清理结果{source.suffix.lower()}"
        )
        temporary_file = temporary_output(output_file)
        workbook.save(temporary_file)
        output_file = Path(publish_output(temporary_file, output_file))
    finally:
        workbook.close()
        if temporary_file:
            Path(temporary_file).unlink(missing_ok=True)

    log_file = available_output_path(output_file.with_name(f"{output_file.stem}_日志.txt"))
    _write_log(
        log_file,
        source,
        options,
        preview,
        output_file,
        final_rows,
        removed_rows,
    )
    return CleanupResult(
        str(output_file),
        str(log_file),
        preview,
        final_rows,
        removed_rows,
    )
