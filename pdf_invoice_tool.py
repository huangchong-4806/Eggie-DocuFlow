import os
import re
import tempfile
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FIELDS = (
    "发票代码",
    "发票号码",
    "开票日期",
    "购买方名称",
    "购买方税号",
    "购买方地址电话",
    "购买方开户行及账号",
    "销售方名称",
    "销售方税号",
    "销售方地址电话",
    "销售方开户行及账号",
    "合计金额（不含税）",
    "合计税额",
    "价税合计（小写）",
    "价税合计（大写）",
    "备注",
)
ITEM_FIELDS = ("项目名称", "规格型号", "单位", "数量", "单价", "金额", "税率", "税额")
VALIDATION_FIELDS = (
    "发票号码",
    "明细行号",
    "校验项",
    "是否异常",
    "异常字段",
    "实际值",
    "应为",
    "差额",
    "说明",
)
MONEY_TOLERANCE = Decimal("0.02")


class ScannedPdfUnsupportedError(RuntimeError):
    def __init__(self, pages):
        page_text = "、".join(map(str, pages))
        super().__init__(f"第 {page_text} 页是扫描图片，当前版本仅支持文本型 PDF 发票。")


@dataclass(frozen=True)
class TextBlock:
    text: str
    page: int
    x0: float = 0
    top: float = 0
    x1: float = 0


@dataclass
class InvoiceItem:
    project_name: str = ""
    specification: str = ""
    unit: str = ""
    quantity: Optional[Decimal] = None
    unit_price: Optional[Decimal] = None
    amount: Optional[Decimal] = None
    tax_rate: Optional[Decimal] = None
    tax_amount: Optional[Decimal] = None


@dataclass(frozen=True)
class ValidationRecord:
    invoice_number: str
    item_row: str
    check: str
    abnormal: bool
    abnormal_fields: str = ""
    actual: Optional[Decimal] = None
    expected: Optional[Decimal] = None
    difference: Optional[Decimal] = None
    note: str = ""


@dataclass
class InvoiceData:
    header: dict = field(default_factory=lambda: {name: "" for name in HEADER_FIELDS})
    items: list = field(default_factory=list)
    validations: list = field(default_factory=list)


@dataclass(frozen=True)
class PdfInvoiceResult:
    output_file: str
    item_count: int
    abnormal_count: int


def _clean_text(value):
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("\u200b", "").replace("\ufeff", "")
    text = re.sub(r"(\d)\s*%", r"\1%", text)
    return re.sub(r"\s+", " ", text).strip(" |丨")


def _compact(value):
    return re.sub(r"[\s:：()（）]", "", _clean_text(value))


def _group_lines(blocks, tolerance=5):
    grouped = []
    for page in sorted({block.page for block in blocks}):
        page_blocks = sorted(
            (block for block in blocks if block.page == page and _clean_text(block.text)),
            key=lambda block: (block.top, block.x0),
        )
        page_lines = []
        for block in page_blocks:
            if not page_lines or abs(page_lines[-1][0].top - block.top) > tolerance:
                page_lines.append([block])
            else:
                page_lines[-1].append(block)
        grouped.extend(sorted(line, key=lambda block: block.x0) for line in page_lines)
    return grouped


def _line_text(line):
    return " ".join(filter(None, (_clean_text(block.text) for block in line)))


def _extract_first(patterns, text):
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return _clean_text(match.group(1))
    return ""


def _extract_party(lines, start_label, end_labels):
    start = next(
        (index for index, line in enumerate(lines) if start_label in _compact(line)),
        None,
    )
    if start is None:
        return {}
    end = next(
        (
            index
            for index in range(start + 1, len(lines))
            if any(label in _compact(lines[index]) for label in end_labels)
        ),
        len(lines),
    )
    result = {"名称": "", "税号": "", "地址电话": "", "开户行及账号": ""}
    label_pattern = re.compile(
        r"(纳税人识别号|统一社会信用代码|"
        r"开户行及账号|开户银行及账号|银行账号|"
        r"地址\s*[、,]?\s*电话|税号|名称)\s*[:：]?",
        re.IGNORECASE,
    )

    def field_name(label):
        compact = re.sub(r"[\s、,]", "", label)
        if compact == "名称":
            return "名称"
        if compact in ("纳税人识别号", "统一社会信用代码", "税号"):
            return "税号"
        if compact == "地址电话":
            return "地址电话"
        return "开户行及账号"

    current_field = None
    for line in lines[start + 1 : end]:
        matches = list(label_pattern.finditer(line))
        if not matches:
            if current_field:
                result[current_field] += _clean_text(line)
            continue
        for index, match in enumerate(matches):
            current_field = field_name(match.group(1))
            value_end = matches[index + 1].start() if index + 1 < len(matches) else len(line)
            result[current_field] += _clean_text(line[match.end() : value_end])
    return result


def _decimal_tokens(text):
    tokens = re.findall(r"[-+]?(?:\d[\d,]*\.?\d*|\.\d+)%?", _clean_text(text))
    return [token for token in tokens if re.search(r"\d", token)]


def _to_decimal(value, percent=False):
    if value in (None, ""):
        return None
    text = _clean_text(value).replace("¥", "").replace("￥", "").replace(",", "")
    is_percent = text.endswith("%")
    text = text.rstrip("%").strip()
    if text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    if percent or is_percent:
        number /= 100
    return number


def _build_item(text_parts, numeric_parts, rate_token):
    numbers = [_to_decimal(token) for token in numeric_parts]
    numbers = [number for number in numbers if number is not None]
    rate = _to_decimal(rate_token, percent=True) if rate_token else None
    if rate is not None:
        if len(numbers) < 4:
            return None
        quantity, unit_price, amount = numbers[-4:-1]
        tax_amount = numbers[-1]
    elif len(numbers) >= 4:
        quantity, unit_price, amount, tax_amount = numbers[-4:]
    elif len(numbers) >= 3:
        quantity, unit_price, amount = numbers[-3:]
        tax_amount = None
    else:
        return None

    text_parts = [_clean_text(part) for part in text_parts if _clean_text(part)]
    item = InvoiceItem(
        project_name=text_parts[0] if text_parts else "",
        specification=text_parts[1] if len(text_parts) > 1 else "",
        unit=" ".join(text_parts[2:]) if len(text_parts) > 2 else "",
        quantity=quantity,
        unit_price=unit_price,
        amount=amount,
        tax_rate=rate,
        tax_amount=tax_amount,
    )
    return item


def _extract_items(lines):
    start = next(
        (
            index
            for index, line in enumerate(lines)
            if "项目名称" in _compact(line)
            and any(label in _compact(line) for label in ("金额", "数量", "税额"))
        ),
        None,
    )
    if start is None:
        return []

    content = []
    for line in lines[start + 1 :]:
        compact = _compact(line)
        if "价税合计" in compact or compact.startswith("合计"):
            break
        if line.strip():
            content.append(line)

    items = []
    text_parts = []
    numeric_parts = []
    rate_token = ""

    def finish():
        nonlocal text_parts, numeric_parts, rate_token
        item = _build_item(text_parts, numeric_parts, rate_token)
        if item and item.amount is not None:
            items.append(item)
        text_parts, numeric_parts, rate_token = [], [], ""

    for line in content:
        tokens = _clean_text(line).split()
        rates = [token for token in tokens if token.endswith("%") and _to_decimal(token, True) is not None]
        numbers = [
            token
            for token in tokens
            if not token.endswith("%") and _to_decimal(token) is not None
        ]
        texts = [token for token in tokens if token not in rates and token not in numbers]

        if texts and numeric_parts:
            finish()
        text_parts.extend(texts)
        numeric_parts.extend(numbers)
        if rates:
            rate_token = rates[-1]
    finish()
    return items


def _column_text(line, minimum, maximum, use_start=False):
    values = []
    for block in line:
        position = block.x0 if use_start else (block.x0 + block.x1) / 2
        if minimum <= position < maximum:
            text = _clean_text(block.text)
            if text:
                values.append(text)
    return "".join(values)


def _positioned_items(blocks):
    items = []
    grouped = _group_lines(blocks, tolerance=3)
    pages = sorted({block.page for block in blocks})
    for page in pages:
        page_lines = [line for line in grouped if line[0].page == page]
        header_index = next(
            (
                index
                for index, line in enumerate(page_lines)
                if "项目名称" in _compact(_line_text(line))
                and "税额" in _compact(_line_text(line))
            ),
            None,
        )
        if header_index is None:
            continue
        subtotal_tops = [
            min(block.top for block in line)
            for line in page_lines
            if _compact(_line_text(line)).startswith("小计")
        ]
        content = []
        for line in page_lines[header_index + 1 :]:
            compact = _compact(_line_text(line))
            if compact.startswith("小计") or compact.startswith("合计"):
                break
            content.append(line)

        numeric_indexes = []
        for index, line in enumerate(content):
            line_top = min(block.top for block in line)
            if any(abs(line_top - top) <= 10 for top in subtotal_tops):
                continue
            amount = _to_decimal(_column_text(line, 385, 450))
            if amount is not None:
                numeric_indexes.append(index)
        if not numeric_indexes:
            continue

        leading_text = "".join(
            _column_text(line, 0, 190, use_start=True)
            for line in content[: numeric_indexes[0]]
        ).strip()
        leading_text = re.split(
            r"\*(?=[\u3400-\u9fff]{1,20}\*)", leading_text, maxsplit=1
        )[0]
        if leading_text and items:
            items[-1].project_name = "".join(
                filter(None, (items[-1].project_name, leading_text))
            )

        for position, line_index in enumerate(numeric_indexes):
            end_index = (
                numeric_indexes[position + 1]
                if position + 1 < len(numeric_indexes)
                else len(content)
            )
            row_lines = content[line_index:end_index]
            first_line = row_lines[0]
            project_name = "".join(
                filter(
                    None,
                    (_column_text(line, 0, 119, use_start=True) for line in row_lines),
                )
            )
            specification = "".join(
                filter(
                    None,
                    (
                        _column_text(line, 119, 190, use_start=True)
                        for line in row_lines
                    ),
                )
            )
            item = InvoiceItem(
                project_name=project_name,
                specification=specification,
                unit=_column_text(first_line, 190, 245),
                quantity=_to_decimal(_column_text(first_line, 245, 310)),
                unit_price=_to_decimal(_column_text(first_line, 310, 385)),
                amount=_to_decimal(_column_text(first_line, 385, 450)),
                tax_rate=_to_decimal(_column_text(first_line, 450, 530), percent=True),
                tax_amount=_to_decimal(_column_text(first_line, 530, float("inf"))),
            )
            if item.amount is not None:
                items.append(item)
    return items


def _parties_from_blocks(blocks):
    first_page = min((block.page for block in blocks), default=1)
    page_blocks = [block for block in blocks if block.page == first_page]
    if not page_blocks:
        return {}, {}
    page_width = max((block.x1 for block in page_blocks), default=600)
    if page_width < 100:
        return {}, {}

    def extract(minimum, maximum):
        result = {}
        for block in page_blocks:
            if not minimum <= block.x0 < maximum:
                continue
            compact = _compact(block.text).replace("、", "").replace("/", "")
            field = None
            if compact == "名称":
                field = "名称"
            elif "纳税人识别号" in compact or "统一社会信用代码" in compact:
                field = "税号"
            elif "地址电话" in compact:
                field = "地址电话"
            elif "开户行及账号" in compact or "开户银行及账号" in compact:
                field = "开户行及账号"
            if not field:
                continue
            values = [
                candidate
                for candidate in page_blocks
                if candidate is not block
                and abs(candidate.top - block.top) <= 4
                and candidate.x0 >= block.x1 - 6
                and candidate.x0 < maximum
            ]
            value = "".join(
                _clean_text(candidate.text)
                for candidate in sorted(values, key=lambda candidate: candidate.x0)
            )
            if value:
                result[field] = value
        return result

    midpoint = page_width / 2
    return extract(20, midpoint), extract(midpoint, page_width + 1)


def _extract_header(lines):
    header = {name: "" for name in HEADER_FIELDS}
    text = "\n".join(lines)
    header["发票代码"] = re.sub(
        r"\s",
        "",
        _extract_first((r"发票代码\s*[:：]?\s*([0-9A-Z ]{8,30})",), text),
    )
    header["发票号码"] = re.sub(
        r"\s",
        "",
        _extract_first((r"发票号码\s*[:：]?\s*([0-9A-Z ]{6,30})",), text),
    )
    header["开票日期"] = _extract_first(
        (
            r"开票日期\s*[:：]?\s*(\d{4}[\-/.\u5e74]\d{1,2}[\-/.\u6708]\d{1,2}\u65e5?)",
            r"开票日期\s*[:：]?\s*(\d{8})",
        ),
        text,
    )

    buyer = _extract_party(lines, "购买方", ("项目名称", "销售方"))
    seller = _extract_party(lines, "销售方", ("收款人", "复核", "开票人"))
    for prefix, party in (("购买方", buyer), ("销售方", seller)):
        header[f"{prefix}名称"] = party.get("名称", "")
        header[f"{prefix}税号"] = party.get("税号", "")
        header[f"{prefix}地址电话"] = party.get("地址电话", "")
        header[f"{prefix}开户行及账号"] = party.get("开户行及账号", "")

    for line in lines:
        compact = _compact(line)
        numbers = _decimal_tokens(line)
        if compact.startswith("合计") and "价税合计" not in compact:
            if numbers:
                header["合计金额（不含税）"] = _to_decimal(numbers[0])
            if len(numbers) > 1:
                header["合计税额"] = _to_decimal(numbers[-1])
        if "价税合计" in compact:
            if numbers:
                header["价税合计（小写）"] = _to_decimal(numbers[-1])
            large = re.search(r"大写[\s:：)）]*([^\d￥¥(（]+)", line)
            if large:
                header["价税合计（大写）"] = _clean_text(large.group(1))

    remark = _extract_first(
        (r"备注\s*[:：]?\s*(.*?)(?=\n销售方|\n收款人|$)",),
        text,
    )
    header["备注"] = remark

    if not header["合计金额（不含税）"]:
        value = _extract_first(
            (r"合计金额(?:（不含税）|\(不含税\))?\s*[:：]?[^数\d-]*([-+]?\d[\d,.]*)",),
            text,
        )
        header["合计金额（不含税）"] = _to_decimal(value)
    if not header["合计税额"]:
        value = _extract_first((r"合计税额\s*[:：]?[^数\d-]*([-+]?\d[\d,.]*)",), text)
        header["合计税额"] = _to_decimal(value)
    if not header["价税合计（小写）"]:
        value = _extract_first(
            (r"价税合计(?:（小写）|\(小写\))?\s*[:：]?[^数\d-]*([-+]?\d[\d,.]*)",),
            text,
        )
        header["价税合计（小写）"] = _to_decimal(value)

    return header


def validate_invoice(invoice):
    records = []
    invoice_number = invoice.header.get("发票号码", "")

    def add(item_row, check, fields, actual, expected, note=""):
        missing = [field for field, value in fields if value is None]
        if missing:
            records.append(
                ValidationRecord(
                    invoice_number,
                    item_row,
                    check,
                    True,
                    "、".join(missing),
                    note="校验所需字段缺失",
                )
            )
            return
        difference = abs(actual - expected)
        records.append(
            ValidationRecord(
                invoice_number,
                item_row,
                check,
                difference > MONEY_TOLERANCE,
                "、".join(field for field, _ in fields) if difference > MONEY_TOLERANCE else "",
                actual,
                expected,
                difference,
                note,
            )
        )

    for index, item in enumerate(invoice.items, 1):
        if item.amount is not None and item.amount < 0 and (
            item.quantity is None or item.unit_price is None
        ):
            records.append(
                ValidationRecord(
                    invoice_number,
                    str(index),
                    "数量 × 单价 ≈ 金额",
                    False,
                    note="折扣/调整行，数量与单价不适用",
                )
            )
        else:
            add(
                str(index),
                "数量 × 单价 ≈ 金额",
                (("数量", item.quantity), ("单价", item.unit_price), ("金额", item.amount)),
                item.quantity * item.unit_price if None not in (item.quantity, item.unit_price) else Decimal(0),
                item.amount or Decimal(0),
            )
        add(
            str(index),
            "金额 × 税率 ≈ 税额",
            (("金额", item.amount), ("税率", item.tax_rate), ("税额", item.tax_amount)),
            item.amount * item.tax_rate if None not in (item.amount, item.tax_rate) else Decimal(0),
            item.tax_amount or Decimal(0),
        )

    amount = invoice.header.get("合计金额（不含税）")
    tax = invoice.header.get("合计税额")
    gross = invoice.header.get("价税合计（小写）")
    add(
        "",
        "金额 + 税额 ≈ 价税合计",
        (("合计金额（不含税）", amount), ("合计税额", tax), ("价税合计（小写）", gross)),
        amount + tax if None not in (amount, tax) else Decimal(0),
        gross or Decimal(0),
    )

    detail_amount = sum((item.amount or Decimal(0) for item in invoice.items), Decimal(0))
    detail_tax = sum((item.tax_amount or Decimal(0) for item in invoice.items), Decimal(0))
    add(
        "",
        "明细金额合计 ≈ 发票合计金额",
        (("明细金额", detail_amount), ("合计金额（不含税）", amount)),
        detail_amount,
        amount or Decimal(0),
    )
    add(
        "",
        "明细税额合计 ≈ 发票合计税额",
        (("明细税额", detail_tax), ("合计税额", tax)),
        detail_tax,
        tax or Decimal(0),
    )
    invoice.validations = records
    return records


def parse_invoice_blocks(blocks):
    grouped = _group_lines(blocks)
    lines = [_line_text(line) for line in grouped]
    lines = [line for line in lines if line]
    invoice_numbers = {
        re.sub(r"\s", "", match)
        for line in lines
        for match in re.findall(
            r"发票号码\s*[:：]?\s*([0-9A-Z ]{6,30})",
            line,
            re.IGNORECASE,
        )
    }
    if len(invoice_numbers) > 1:
        raise ValueError("检测到多张发票，请拆分为每个 PDF 一张发票后再处理。")
    header = _extract_header(lines)
    buyer, seller = _parties_from_blocks(blocks)
    for prefix, party in (("购买方", buyer), ("销售方", seller)):
        for suffix in ("名称", "税号", "地址电话", "开户行及账号"):
            if party.get(suffix) and not header[f"{prefix}{suffix}"]:
                header[f"{prefix}{suffix}"] = party[suffix]
    items = _positioned_items(blocks)
    if not items:
        page_lines = {}
        for line in grouped:
            page_lines.setdefault(line[0].page, []).append(_line_text(line))
        items = [
            item
            for page in sorted(page_lines)
            for item in _extract_items(page_lines[page])
        ]
    invoice = InvoiceData(header=header, items=items)
    if not invoice.items:
        raise ValueError("未能识别包含“金额”的发票明细，未生成 Excel。")
    validate_invoice(invoice)
    return invoice


def _pdf_text_blocks(pdf_file, progress_callback=None):
    try:
        import pdfplumber
    except ImportError as error:
        raise RuntimeError("缺少 PDF 文本解析组件 pdfplumber。") from error

    blocks = []
    scanned_pages = []
    with pdfplumber.open(pdf_file) as document:
        total = len(document.pages)
        if total == 0:
            raise ValueError("PDF 中没有可处理页面。")
        if total > 100:
            raise ValueError("PDF 超过 100 页，请拆分后再处理。")
        for page_number, page in enumerate(document.pages, 1):
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
            visible_text = "".join(str(word.get("text", "")) for word in words)
            page_area = max(float(page.width) * float(page.height), 1)
            has_page_image = any(
                float(image.get("width", 0)) * float(image.get("height", 0))
                >= page_area * 0.25
                for image in page.images
            )
            if len(re.sub(r"\s", "", visible_text)) < 20 and has_page_image:
                scanned_pages.append(page_number)
            else:
                blocks.extend(
                    TextBlock(
                        text=word.get("text", ""),
                        page=page_number,
                        x0=float(word.get("x0", 0)),
                        top=float(word.get("top", 0)),
                        x1=float(word.get("x1", 0)),
                    )
                    for word in words
                )
            if progress_callback:
                progress_callback(page_number, total, f"正在读取第 {page_number} 页")
    return blocks, scanned_pages


def extract_invoice(pdf_file, progress_callback=None):
    pdf_file = os.path.abspath(pdf_file)
    if Path(pdf_file).suffix.lower() != ".pdf":
        raise ValueError("只支持 PDF 文件。")
    if not os.path.isfile(pdf_file):
        raise FileNotFoundError("PDF 文件不存在。")
    blocks, scanned_pages = _pdf_text_blocks(pdf_file, progress_callback)
    if scanned_pages:
        raise ScannedPdfUnsupportedError(scanned_pages)
    return parse_invoice_blocks(blocks)


def _excel_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if text.startswith(("=", "+", "-", "@")):
            return "'" + text
        return text
    return value


def _date_value(value):
    if not value:
        return ""
    matched = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", str(value))
    if matched:
        try:
            return datetime(
                int(matched.group(1)),
                int(matched.group(2)),
                int(matched.group(3)),
            ).date()
        except ValueError:
            pass
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%Y%m%d").date()
        except ValueError:
            pass
    return _excel_value(value)


def _style_sheet(sheet, widths):
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 26
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def write_invoice_workbook(invoice, output_file, overwrite=False):
    output_file = os.path.abspath(output_file)
    if Path(output_file).suffix.lower() != ".xlsx":
        raise ValueError("输出文件必须为 .xlsx 格式。")
    output_parent = os.path.dirname(output_file)
    if not os.path.isdir(output_parent):
        raise ValueError("输出文件夹不存在。")
    if os.path.exists(output_file) and not overwrite:
        raise FileExistsError("输出文件已存在，请选择新的保存位置。")

    workbook = Workbook()
    header_sheet = workbook.active
    header_sheet.title = "发票头信息"
    item_sheet = workbook.create_sheet("明细表")
    validation_sheet = workbook.create_sheet("校验结果")

    header_sheet.append(HEADER_FIELDS)
    header_sheet.append(
        [
            _date_value(invoice.header[field])
            if field == "开票日期"
            else _excel_value(invoice.header[field])
            for field in HEADER_FIELDS
        ]
    )
    for column in (12, 13, 14):
        header_sheet.cell(2, column).number_format = '#,##0.00;[Red]-#,##0.00;"-"'
    for column in (1, 2, 5, 9):
        header_sheet.cell(2, column).number_format = "@"
        header_sheet.cell(2, column).quotePrefix = True
    header_sheet.cell(2, 3).number_format = "yyyy-mm-dd"

    item_sheet.append(ITEM_FIELDS)
    for item in invoice.items:
        item_sheet.append(
            [
                _excel_value(item.project_name),
                _excel_value(item.specification),
                _excel_value(item.unit),
                _excel_value(item.quantity),
                _excel_value(item.unit_price),
                _excel_value(item.amount),
                _excel_value(item.tax_rate),
                _excel_value(item.tax_amount),
            ]
        )
    for row in range(2, item_sheet.max_row + 1):
        for column in (4, 5, 6, 8):
            item_sheet.cell(row, column).number_format = '#,##0.00####;[Red]-#,##0.00####;"-"'
        item_sheet.cell(row, 7).number_format = "0.00%"

    validation_sheet.append(VALIDATION_FIELDS)
    for record in invoice.validations:
        validation_sheet.append(
            [
                _excel_value(record.invoice_number),
                record.item_row,
                record.check,
                "异常" if record.abnormal else "正常",
                record.abnormal_fields,
                _excel_value(record.actual),
                _excel_value(record.expected),
                _excel_value(record.difference),
                record.note,
            ]
        )
        status_cell = validation_sheet.cell(validation_sheet.max_row, 4)
        status_cell.fill = PatternFill(
            "solid",
            fgColor="F4CCCC" if record.abnormal else "D9EAD3",
        )
        status_cell.font = Font(color="9C0006" if record.abnormal else "274E13", bold=True)
    for row in range(2, validation_sheet.max_row + 1):
        validation_sheet.cell(row, 1).number_format = "@"
        validation_sheet.cell(row, 1).quotePrefix = True
        for column in (6, 7, 8):
            validation_sheet.cell(row, column).number_format = '#,##0.00;[Red]-#,##0.00;"-"'

    _style_sheet(
        header_sheet,
        (15, 18, 13, 24, 22, 28, 30, 24, 22, 28, 30, 18, 14, 18, 20, 32),
    )
    _style_sheet(item_sheet, (80, 24, 10, 12, 14, 14, 10, 14))
    _style_sheet(
        validation_sheet,
        (18, 11, 30, 12, 30, 16, 16, 14, 28),
    )

    file_descriptor, temporary_file = tempfile.mkstemp(
        prefix=f".{Path(output_file).stem}-",
        suffix=".xlsx",
        dir=output_parent,
    )
    os.close(file_descriptor)
    try:
        workbook.save(temporary_file)
        os.chmod(temporary_file, 0o644)
        if overwrite:
            os.replace(temporary_file, output_file)
        else:
            os.link(temporary_file, output_file)
            os.unlink(temporary_file)
    except Exception:
        try:
            os.unlink(temporary_file)
        except FileNotFoundError:
            pass
        raise
    finally:
        workbook.close()
    return output_file


def convert_invoice_pdf(pdf_file, output_file, progress_callback=None, overwrite=False):
    invoice = extract_invoice(pdf_file, progress_callback)
    write_invoice_workbook(invoice, output_file, overwrite=overwrite)
    return PdfInvoiceResult(
        output_file=os.path.abspath(output_file),
        item_count=len(invoice.items),
        abnormal_count=sum(record.abnormal for record in invoice.validations),
    )


def convert_invoice_pdfs(pdf_files, output_folder, progress_callback=None):
    output_folder = os.path.abspath(output_folder)
    if not os.path.isdir(output_folder):
        raise ValueError("Excel 保存文件夹不存在。")

    results = []
    failures = []
    total = len(pdf_files)
    for index, pdf_file in enumerate(pdf_files, 1):
        if progress_callback:
            progress_callback(index - 1, total, f"正在解析：{Path(pdf_file).name}")
        output_file = Path(output_folder) / f"{Path(pdf_file).stem}_发票结构化.xlsx"
        number = 1
        while output_file.exists():
            output_file = output_file.with_name(
                f"{Path(pdf_file).stem}_发票结构化_{number}.xlsx"
            )
            number += 1
        try:
            results.append(convert_invoice_pdf(pdf_file, output_file))
        except Exception as error:
            failures.append((os.path.abspath(pdf_file), str(error)))

    if progress_callback:
        progress_callback(total, total, "批量解析完成")
    return results, failures
