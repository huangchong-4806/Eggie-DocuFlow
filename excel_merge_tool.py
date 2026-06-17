import os
import posixpath
import re
import time
import warnings
from copy import copy
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}

_ATTRIBUTE_PATTERN = re.compile(
    rb"([A-Za-z_:][A-Za-z0-9_.:-]*)\s*=\s*(?:\"([^\"]*)\"|'([^']*)')"
)
_DIMENSION_PATTERN = re.compile(
    rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?dimension\b[^>]*"
    rb"\bref\s*=\s*(?:\"([^\"]+)\"|'([^']+)')"
)
_ROW_PATTERN = re.compile(
    rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?row\b[^>]*"
    rb"\br\s*=\s*(?:\"(\d+)\"|'(\d+)')"
)
_COLUMN_PATTERN = re.compile(
    rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?col\b[^>]*/?>"
)
_MERGED_CELL_PATTERN = re.compile(
    rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?mergeCell\b[^>]*"
    rb"\bref\s*=\s*(?:\"([^\"]+)\"|'([^']+)')"
)


@dataclass(frozen=True)
class WorkbookMetadata:
    row_count: int
    column_widths: tuple
    merged_ranges: tuple
    worksheet_path: str


@dataclass(frozen=True)
class SplitWorkbookResult:
    output_files: tuple
    output_folder: str
    total_rows: int
    header_rows: int
    data_rows: int
    elapsed_seconds: float

    @property
    def file_count(self):
        return len(self.output_files)

    @property
    def average_seconds_per_file(self):
        if not self.output_files:
            return 0
        return self.elapsed_seconds / len(self.output_files)


def discover_excel_files(folder):
    excel_files = []

    for root, directory_names, filenames in os.walk(folder):
        directory_names[:] = sorted(
            (
                name
                for name in directory_names
                if not name.startswith(".") and name != "__MACOSX"
            ),
            key=str.casefold,
        )

        for filename in sorted(filenames, key=str.casefold):
            if filename.startswith(("~$", ".")):
                continue
            if Path(filename).suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue

            path = Path(root, filename)
            if path.is_file():
                excel_files.append(str(path.resolve()))

    return excel_files


def _local_name(tag):
    return tag.rsplit("}", 1)[-1]


def _relationship_id(element):
    for attribute_name, value in element.attrib.items():
        if attribute_name == "id" or attribute_name.endswith("}id"):
            return value
    return None


def _active_worksheet_path(archive):
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    active_tab = 0

    for element in workbook_root.iter():
        if _local_name(element.tag) == "workbookView":
            try:
                active_tab = int(element.attrib.get("activeTab", 0))
            except (TypeError, ValueError):
                active_tab = 0
            break

    sheets = [
        element
        for element in workbook_root.iter()
        if _local_name(element.tag) == "sheet"
    ]
    if not sheets:
        raise ValueError("Excel 文件中没有工作表。")

    active_tab = max(0, min(active_tab, len(sheets) - 1))
    relationship_id = _relationship_id(sheets[active_tab])
    if not relationship_id:
        raise ValueError("无法确定 Excel 活动工作表。")

    relationships_root = ElementTree.fromstring(
        archive.read("xl/_rels/workbook.xml.rels")
    )
    target = None
    for relationship in relationships_root:
        if relationship.attrib.get("Id") == relationship_id:
            target = relationship.attrib.get("Target")
            break

    if not target:
        raise ValueError("无法读取 Excel 活动工作表路径。")

    if target.startswith("/"):
        worksheet_path = target.lstrip("/")
    else:
        worksheet_path = posixpath.normpath(posixpath.join("xl", target))

    if worksheet_path not in archive.namelist():
        raise ValueError("Excel 活动工作表文件不存在。")
    return worksheet_path


def _attribute_values(tag):
    values = {}
    for match in _ATTRIBUTE_PATTERN.finditer(tag):
        name = match.group(1).decode("ascii", errors="ignore").split(":")[-1]
        raw_value = match.group(2) or match.group(3) or b""
        values[name] = raw_value.decode("utf-8", errors="replace")
    return values


def _matched_value(match):
    return (match.group(1) or match.group(2)).decode(
        "utf-8",
        errors="replace",
    )


def _dimension_row_count(dimension_reference):
    if not dimension_reference:
        return 0
    try:
        _, _, _, max_row = range_boundaries(dimension_reference)
    except (TypeError, ValueError):
        return 0
    return max_row or 0


def _scan_worksheet_metadata(archive, worksheet_path):
    dimension_reference = None
    max_row = 0
    column_widths = {}
    merged_ranges = set()
    tail = b""

    with archive.open(worksheet_path) as worksheet_file:
        while True:
            chunk = worksheet_file.read(1024 * 1024)
            if not chunk:
                break

            data = tail + chunk
            if dimension_reference is None:
                match = _DIMENSION_PATTERN.search(data)
                if match:
                    dimension_reference = _matched_value(match)

            for match in _ROW_PATTERN.finditer(data):
                row_number = int(match.group(1) or match.group(2))
                max_row = max(max_row, row_number)

            for match in _COLUMN_PATTERN.finditer(data):
                attributes = _attribute_values(match.group(0))
                try:
                    min_column = int(attributes["min"])
                    max_column = int(attributes["max"])
                    width = float(attributes["width"])
                except (KeyError, TypeError, ValueError):
                    continue
                column_widths[(min_column, max_column)] = width

            for match in _MERGED_CELL_PATTERN.finditer(data):
                merged_ranges.add(_matched_value(match))

            tail = data[-4096:]

    row_count = max(max_row, _dimension_row_count(dimension_reference))
    if row_count == 0 and dimension_reference:
        row_count = 1

    return WorkbookMetadata(
        row_count=row_count,
        column_widths=tuple(
            (min_column, max_column, width)
            for (min_column, max_column), width in sorted(column_widths.items())
        ),
        merged_ranges=tuple(sorted(merged_ranges)),
        worksheet_path=worksheet_path,
    )


@lru_cache(maxsize=128)
def _cached_workbook_metadata(filename, modified_time_ns, size):
    del modified_time_ns, size
    with ZipFile(filename) as archive:
        worksheet_path = _active_worksheet_path(archive)
        return _scan_worksheet_metadata(archive, worksheet_path)


def get_workbook_metadata(filename):
    absolute_filename = os.path.abspath(filename)
    file_status = os.stat(absolute_filename)
    return _cached_workbook_metadata(
        absolute_filename,
        file_status.st_mtime_ns,
        file_status.st_size,
    )


def _style_identifier(source_cell):
    if hasattr(source_cell, "_style_id"):
        return source_cell._style_id
    return source_cell.style_id


def copy_cell_style(source_cell, target_cell, style_cache=None):
    if not source_cell.has_style:
        return

    style_identifier = _style_identifier(source_cell)
    if style_cache is not None and style_identifier in style_cache:
        target_cell._style = copy(style_cache[style_identifier])
        return

    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)

    if style_cache is not None:
        style_cache[style_identifier] = copy(target_cell._style)


def _translated_cell_value(source_cell, target_coordinate):
    value = source_cell.value
    if source_cell.data_type != "f" or not isinstance(value, str):
        return value

    try:
        return Translator(
            value,
            origin=source_cell.coordinate,
        ).translate_formula(target_coordinate)
    except Exception:
        return value


def copy_cell_value(source_cell, target_cell):
    value = _translated_cell_value(source_cell, target_cell.coordinate)
    target_cell.value = value

    # Keep formula-looking source text as text instead of converting it to a formula.
    if (
        source_cell.data_type == "s"
        and isinstance(value, str)
        and value.startswith("=")
    ):
        target_cell.data_type = "s"


def format_file_size(size):
    value = float(size)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            if unit == "B":
                return f"{int(value)} {unit}"
            return f"{value:.1f} {unit}"
        value /= 1024


def get_file_info(filename):
    file_size = os.path.getsize(filename)
    metadata = get_workbook_metadata(filename)
    return {
        "size": format_file_size(file_size),
        "rows": metadata.row_count,
    }


def _apply_column_widths(output_sheet, metadata):
    for min_column, max_column, width in metadata.column_widths:
        for column_number in range(min_column, max_column + 1):
            output_sheet.column_dimensions[
                get_column_letter(column_number)
            ].width = width


def _contains_relevant_merged_cells(metadata, start_row):
    for merged_range in metadata.merged_ranges:
        try:
            _, min_row, _, _ = range_boundaries(merged_range)
        except (TypeError, ValueError):
            return True
        if min_row >= start_row:
            return True
    return False


def _configure_formula_calculation(workbook):
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True


def _copy_column_dimensions(source_sheet, output_sheet):
    for column_key, source_dimension in source_sheet.column_dimensions.items():
        target_dimension = output_sheet.column_dimensions[column_key]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden


def _copy_row_dimension(source_sheet, output_sheet, source_row, target_row):
    source_dimension = source_sheet.row_dimensions[source_row]
    if source_dimension.height is None and not source_dimension.hidden:
        return

    target_dimension = output_sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden


def _copy_auto_filter(source_sheet, output_sheet, output_max_row):
    if not source_sheet.auto_filter.ref:
        return

    try:
        min_column, min_row, max_column, _ = range_boundaries(
            source_sheet.auto_filter.ref
        )
    except (TypeError, ValueError):
        output_sheet.auto_filter.ref = source_sheet.auto_filter.ref
        return

    output_sheet.auto_filter.ref = (
        f"{get_column_letter(min_column)}{min_row}:"
        f"{get_column_letter(max_column)}{max(output_max_row, min_row)}"
    )


def _cached_cell_style_parts(source_cell, style_cache):
    if not source_cell.has_style:
        return None

    style_identifier = _style_identifier(source_cell)
    if style_identifier not in style_cache:
        style_cache[style_identifier] = (
            copy(source_cell.font),
            copy(source_cell.fill),
            copy(source_cell.border),
            copy(source_cell.alignment),
            source_cell.number_format,
            copy(source_cell.protection),
        )
    return style_cache[style_identifier]


def _copy_split_cell(source_cell, target_cell, style_cache):
    value = source_cell.value
    if source_cell.data_type == "f" and isinstance(value, str):
        value = _translated_cell_value(source_cell, target_cell.coordinate)

    target_cell.value = value
    if (
        source_cell.data_type == "s"
        and isinstance(value, str)
        and value.startswith("=")
    ):
        target_cell.data_type = "s"

    cached_style_parts = _cached_cell_style_parts(source_cell, style_cache)
    if cached_style_parts is not None:
        (
            target_cell.font,
            target_cell.fill,
            target_cell.border,
            target_cell.alignment,
            target_cell.number_format,
            target_cell.protection,
        ) = cached_style_parts


def _copy_worksheet_row(
    source_sheet,
    output_sheet,
    source_row_number,
    source_row,
    target_row,
    style_cache,
):
    _copy_row_dimension(source_sheet, output_sheet, source_row_number, target_row)
    for source_cell in source_row:
        target_cell = output_sheet.cell(
            row=target_row,
            column=source_cell.column,
        )
        _copy_split_cell(source_cell, target_cell, style_cache)


def _prepare_split_merged_ranges(source_sheet, header_rows):
    header_ranges = []
    data_ranges = []

    for merged_range in source_sheet.merged_cells.ranges:
        min_column, min_row, max_column, max_row = merged_range.bounds
        bounds = (min_column, min_row, max_column, max_row)
        if min_row >= 1 and max_row <= header_rows:
            header_ranges.append(bounds)
        elif min_row > header_rows:
            data_ranges.append(bounds)

    return header_ranges, data_ranges


def _copy_prepared_split_merged_cells(
    output_sheet,
    header_ranges,
    data_ranges,
    header_rows,
    chunk_start_row,
    chunk_end_row,
):
    for min_column, min_row, max_column, max_row in header_ranges:
        output_sheet.merge_cells(
            start_row=min_row,
            start_column=min_column,
            end_row=max_row,
            end_column=max_column,
        )

    row_offset = header_rows + 1 - chunk_start_row
    for min_column, min_row, max_column, max_row in data_ranges:
        if min_row < chunk_start_row or max_row > chunk_end_row:
            continue

        output_sheet.merge_cells(
            start_row=min_row + row_offset,
            start_column=min_column,
            end_row=max_row + row_offset,
            end_column=max_column,
        )


def _create_split_output_folder(parent_folder, source_file):
    parent_folder = Path(parent_folder)
    source_stem = Path(source_file).stem
    base_name = f"{source_stem}_拆分结果"
    candidate = parent_folder / base_name
    suffix_number = 1

    while True:
        try:
            candidate.mkdir()
            return candidate
        except FileExistsError:
            candidate = parent_folder / f"{base_name}_{suffix_number}"
            suffix_number += 1


def _split_output_path(output_folder, source_file, part_number):
    output_folder = Path(output_folder)
    source_stem = Path(source_file).stem
    return output_folder / f"{source_stem}_拆分{part_number:03d}.xlsx"


def split_workbook_by_rows(
    source_file,
    output_folder,
    rows_per_file,
    header_rows=0,
    progress_callback=None,
):
    if rows_per_file < 1:
        raise ValueError("每个文件的数据行数必须大于 0。")
    if header_rows < 0:
        raise ValueError("表头行数不能小于 0。")

    source_file = os.path.abspath(source_file)
    output_folder = os.path.abspath(output_folder)

    if Path(source_file).suffix.lower() != ".xlsx":
        raise ValueError("拆分工具只支持 .xlsx 格式的 Excel 文件。")
    if not os.path.isfile(source_file):
        raise ValueError("源 Excel 文件不存在。")
    if not os.path.isdir(output_folder):
        raise ValueError("输出文件夹不存在。")

    start_time = time.perf_counter()
    output_files = []

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(
            source_file,
            data_only=False,
            keep_links=False,
        )

    try:
        source_sheet = workbook.active
        max_row = source_sheet.max_row
        if header_rows > max_row:
            raise ValueError("表头行数不能大于总行数。")

        data_start_row = header_rows + 1
        data_rows = max_row - header_rows
        if data_rows == 0:
            raise ValueError("源文件中没有可拆分数据。")

        total_parts = (data_rows + rows_per_file - 1) // rows_per_file
        split_output_folder = _create_split_output_folder(
            output_folder,
            source_file,
        )
        if header_rows:
            header_rows_cache = list(
                source_sheet.iter_rows(
                    min_row=1,
                    max_row=min(header_rows, max_row),
                )
            )
        else:
            header_rows_cache = []
        header_ranges, data_merged_ranges = _prepare_split_merged_ranges(
            source_sheet,
            header_rows,
        )

        for part_index in range(total_parts):
            chunk_start_row = data_start_row + part_index * rows_per_file
            chunk_end_row = min(max_row, chunk_start_row + rows_per_file - 1)
            output_workbook = Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = source_sheet.title
            _configure_formula_calculation(output_workbook)
            _copy_column_dimensions(source_sheet, output_sheet)
            if source_sheet.freeze_panes:
                output_sheet.freeze_panes = source_sheet.freeze_panes

            style_cache = {}
            target_row = 1
            for source_row_number, source_row in enumerate(
                header_rows_cache,
                start=1,
            ):
                _copy_worksheet_row(
                    source_sheet,
                    output_sheet,
                    source_row_number,
                    source_row,
                    target_row,
                    style_cache,
                )
                target_row += 1

            for source_row_number, source_row in enumerate(
                source_sheet.iter_rows(
                    min_row=chunk_start_row,
                    max_row=chunk_end_row,
                ),
                start=chunk_start_row,
            ):
                _copy_worksheet_row(
                    source_sheet,
                    output_sheet,
                    source_row_number,
                    source_row,
                    target_row,
                    style_cache,
                )
                target_row += 1

            output_max_row = target_row - 1
            _copy_prepared_split_merged_cells(
                output_sheet,
                header_ranges,
                data_merged_ranges,
                header_rows,
                chunk_start_row,
                chunk_end_row,
            )
            _copy_auto_filter(source_sheet, output_sheet, output_max_row)

            output_path = _split_output_path(
                split_output_folder,
                source_file,
                part_index + 1,
            )
            try:
                output_workbook.save(output_path)
            finally:
                output_workbook.close()

            output_files.append(str(output_path))
            if progress_callback:
                progress_callback(
                    part_index + 1,
                    total_parts,
                    os.path.basename(output_path),
                )
    finally:
        workbook.close()

    elapsed_seconds = time.perf_counter() - start_time
    return SplitWorkbookResult(
        output_files=tuple(output_files),
        output_folder=str(split_output_folder),
        total_rows=max_row,
        header_rows=header_rows,
        data_rows=data_rows,
        elapsed_seconds=elapsed_seconds,
    )


def _stream_output_cell(
    source_cell,
    output_sheet,
    target_row,
    style_cache,
):
    value = source_cell.value
    has_style = getattr(source_cell, "has_style", False)
    formula_text = (
        getattr(source_cell, "data_type", None) == "f"
        and isinstance(value, str)
    )
    formula_looking_text = (
        getattr(source_cell, "data_type", None) == "s"
        and isinstance(value, str)
        and value.startswith("=")
    )

    if not has_style and not formula_text and not formula_looking_text:
        return value

    target_cell = WriteOnlyCell(output_sheet, value=None)
    target_coordinate = (
        f"{get_column_letter(source_cell.column)}{target_row}"
    )
    target_cell.value = _translated_cell_value(
        source_cell,
        target_coordinate,
    )

    if formula_looking_text:
        target_cell.data_type = "s"

    copy_cell_style(source_cell, target_cell, style_cache)
    return target_cell


def _row_requires_special_handling(source_row):
    for source_cell in source_row:
        if getattr(source_cell, "has_style", False):
            return True

        value = source_cell.value
        data_type = getattr(source_cell, "data_type", None)
        if data_type == "f":
            return True
        if (
            data_type == "s"
            and isinstance(value, str)
            and value.startswith("=")
        ):
            return True
    return False


def _build_streaming_workbook(
    files,
    metadata_by_file,
    output_file,
    skip_rows,
    progress_callback,
):
    output_workbook = Workbook(write_only=True)
    output_sheet = output_workbook.create_sheet("合并结果")
    _configure_formula_calculation(output_workbook)
    _apply_column_widths(output_sheet, metadata_by_file[0])
    target_row = 1

    try:
        for file_index, filename in enumerate(files):
            metadata = metadata_by_file[file_index]
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                workbook = load_workbook(
                    filename,
                    read_only=True,
                    data_only=False,
                    keep_links=False,
                )

            try:
                worksheet = workbook.active
                if metadata.row_count > worksheet.max_row:
                    worksheet.reset_dimensions()

                start_row = skip_rows + 1 if file_index > 0 else 1
                style_cache = {}
                for source_row in worksheet.iter_rows(min_row=start_row):
                    if _row_requires_special_handling(source_row):
                        output_row = [
                            _stream_output_cell(
                                source_cell,
                                output_sheet,
                                target_row,
                                style_cache,
                            )
                            for source_cell in source_row
                        ]
                    else:
                        output_row = [
                            source_cell.value
                            for source_cell in source_row
                        ]

                    output_sheet.append(output_row)
                    target_row += 1
            finally:
                workbook.close()

            if progress_callback:
                progress_callback(file_index + 1, os.path.basename(filename))

        output_workbook.save(output_file)
    finally:
        output_workbook.close()


def _build_compatibility_workbook(
    files,
    metadata_by_file,
    output_file,
    skip_rows,
    keep_merged_cells,
    progress_callback,
):
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "合并结果"
    _configure_formula_calculation(output_workbook)
    _apply_column_widths(output_sheet, metadata_by_file[0])
    current_row = 1

    try:
        for file_index, filename in enumerate(files):
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                workbook = load_workbook(
                    filename,
                    data_only=False,
                    keep_links=False,
                )

            try:
                worksheet = workbook.active
                start_row = skip_rows + 1 if file_index > 0 else 1
                row_offset = current_row - start_row
                style_cache = {}
                copied_rows = 0

                for source_row in worksheet.iter_rows(min_row=start_row):
                    copied_rows += 1
                    for source_cell in source_row:
                        target_cell = output_sheet.cell(
                            row=source_cell.row + row_offset,
                            column=source_cell.column,
                        )
                        copy_cell_value(source_cell, target_cell)
                        copy_cell_style(
                            source_cell,
                            target_cell,
                            style_cache,
                        )

                if keep_merged_cells:
                    for merged_range in worksheet.merged_cells.ranges:
                        min_col, min_row, max_col, max_row = (
                            merged_range.bounds
                        )
                        if min_row < start_row:
                            continue

                        output_sheet.merge_cells(
                            start_row=min_row + row_offset,
                            start_column=min_col,
                            end_row=max_row + row_offset,
                            end_column=max_col,
                        )

                current_row += copied_rows
            finally:
                workbook.close()

            if progress_callback:
                progress_callback(file_index + 1, os.path.basename(filename))

        output_workbook.save(output_file)
    finally:
        output_workbook.close()


def build_merged_workbook(
    files,
    output_file,
    skip_rows=1,
    keep_merged_cells=True,
    progress_callback=None,
):
    if not files:
        raise ValueError("至少需要选择一个 Excel 文件。")

    metadata_by_file = [
        get_workbook_metadata(filename)
        for filename in files
    ]
    requires_compatibility_mode = False

    if keep_merged_cells:
        for file_index, metadata in enumerate(metadata_by_file):
            start_row = skip_rows + 1 if file_index > 0 else 1
            if _contains_relevant_merged_cells(metadata, start_row):
                requires_compatibility_mode = True
                break

    if requires_compatibility_mode:
        _build_compatibility_workbook(
            files,
            metadata_by_file,
            output_file,
            skip_rows,
            keep_merged_cells,
            progress_callback,
        )
        return

    _build_streaming_workbook(
        files,
        metadata_by_file,
        output_file,
        skip_rows,
        progress_callback,
    )
