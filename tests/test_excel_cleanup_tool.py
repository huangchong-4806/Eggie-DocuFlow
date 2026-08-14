import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from excel_cleanup_tool import (
    CleanupOptions,
    clean_workbook,
    preview_cleanup,
    workbook_sheet_names,
)


class ExcelCleanupToolTests(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.source = self.root / "source.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "数据"
        sheet.append(["姓名", "数量", "日期", "计算"])
        sheet.append([" Alice  ", 1, date(2026, 8, 1), "=B2*2"])
        sheet.append(["Alice", 1, date(2026, 8, 1), "=B3*2"])
        sheet.append([None, None, None, None])
        sheet.append(["Bob   Smith", 2.5, date(2026, 8, 2), "=B5*2"])
        workbook.create_sheet("保留")
        workbook.save(self.source)

    def tearDown(self):
        self.temporary_directory.cleanup()

    def options(self, **values):
        defaults = dict(
            sheet_name="数据",
            remove_empty_rows=True,
            trim_whitespace=True,
            deduplicate=True,
            duplicate_columns=(1, 2, 3),
            normalize_dates=True,
            normalize_numbers=True,
        )
        defaults.update(values)
        return CleanupOptions(**defaults)

    def test_preview_reports_changes_without_modifying_source(self):
        preview = preview_cleanup(self.source, self.options())

        self.assertEqual(workbook_sheet_names(self.source), ("数据", "保留"))
        self.assertEqual(preview.blank_rows, 1)
        self.assertEqual(preview.duplicate_rows, 1)
        self.assertEqual(preview.whitespace_cells, 2)
        self.assertEqual(preview.formula_cells, 3)

        workbook = load_workbook(self.source, data_only=False)
        try:
            self.assertEqual(workbook["数据"]["A2"].value, " Alice  ")
            self.assertEqual(workbook["数据"].max_row, 5)
        finally:
            workbook.close()

    def test_preview_reads_workbook_without_saved_dimension(self):
        original_init = ReadOnlyWorksheet.__init__

        def init_without_dimension(sheet, *args, **kwargs):
            original_init(sheet, *args, **kwargs)
            sheet._max_row = None
            sheet._max_column = None

        with patch.object(ReadOnlyWorksheet, "__init__", init_without_dimension):
            preview = preview_cleanup(self.source, self.options())

        self.assertEqual(preview.original_rows, 4)
        self.assertEqual(preview.original_columns, 4)

    def test_cleanup_saves_copy_keeps_other_sheet_and_writes_log(self):
        result = clean_workbook(self.source, self.root / "output", self.options())

        workbook = load_workbook(result.output_file, data_only=False)
        try:
            sheet = workbook["数据"]
            self.assertEqual(sheet.max_row, 3)
            self.assertEqual(sheet["A2"].value, "Alice")
            self.assertEqual(sheet["A3"].value, "Bob Smith")
            self.assertEqual(sheet["B3"].number_format, "#,##0.00")
            self.assertEqual(sheet["C2"].number_format, "yyyy-mm-dd")
            self.assertEqual(sheet["D2"].value, "=B2*2")
            self.assertEqual(sheet["D3"].value, "=B3*2")
            self.assertIn("保留", workbook.sheetnames)
        finally:
            workbook.close()

        self.assertEqual(load_workbook(self.source)["数据"].max_row, 5)
        self.assertEqual(result.final_rows, 2)
        log_text = Path(result.log_file).read_text(encoding="utf-8")
        self.assertIn("blank_rows=1", log_text)
        self.assertIn("duplicate_rows=1", log_text)
        self.assertIn(f"output_file={result.output_file}", log_text)

    def test_cleanup_can_keep_blank_and_duplicate_rows(self):
        options = self.options(
            remove_empty_rows=False,
            deduplicate=False,
            trim_whitespace=False,
            normalize_dates=False,
            normalize_numbers=False,
        )
        result = clean_workbook(self.source, self.root / "output", options)

        workbook = load_workbook(result.output_file)
        try:
            self.assertEqual(workbook["数据"].max_row, 5)
            self.assertEqual(workbook["数据"]["A2"].value, " Alice  ")
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
