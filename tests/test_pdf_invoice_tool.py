import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch
from types import SimpleNamespace
from zipfile import ZipFile

from openpyxl import load_workbook

from pdf_invoice_tool import (
    HEADER_FIELDS,
    ITEM_FIELDS,
    InvoiceData,
    InvoiceItem,
    PdfInvoiceResult,
    ScannedPdfUnsupportedError,
    TextBlock,
    _extract_party,
    _parties_from_blocks,
    _pdf_text_blocks,
    _positioned_items,
    convert_invoice_pdf,
    convert_invoice_pdfs,
    extract_invoice,
    parse_invoice_blocks,
    validate_invoice,
    write_invoice_ledger,
    write_invoice_workbook,
)


def invoice_blocks(item_lines=None):
    lines = [
        "增值税电子普通发票",
        "发票代码: 044031900111",
        "发票号码: 12345678",
        "开票日期: 2026年6月23日",
        "购买方",
        "名称: 购方公司",
        "纳税人识别号: 91310000123456789X",
        "地址、电话: 上海市 021-12345678",
        "开户行及账号: 测试银行 123456789",
        "项目名称 规格型号 单位 数量 单价 金额 税率 税额",
        *(item_lines or ["服务费 1 100.00 100.00 13% 13.00"]),
        "合计 ¥100.00 ¥13.00",
        "价税合计（大写） 壹佰壹拾叁元整 （小写） ¥113.00",
        "备注: 财务入账",
        "销售方",
        "名称: 销方公司",
        "纳税人识别号: 91310000987654321X",
        "地址、电话: 北京市 010-12345678",
        "开户行及账号: 另一银行 987654321",
    ]
    return [
        TextBlock(text=line, page=1, top=index * 20)
        for index, line in enumerate(lines)
    ]


class PdfInvoiceToolTests(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_directory.name)

    def tearDown(self):
        self.temp_directory.cleanup()

    def test_parser_extracts_structured_invoice_and_validates_it(self):
        invoice = parse_invoice_blocks(invoice_blocks())

        self.assertEqual(invoice.header["发票号码"], "12345678")
        self.assertEqual(invoice.header["购买方名称"], "购方公司")
        self.assertEqual(invoice.header["销售方税号"], "91310000987654321X")
        self.assertEqual(invoice.header["价税合计（小写）"], Decimal("113.00"))
        self.assertEqual(invoice.header["价税合计（大写）"], "壹佰壹拾叁元整")
        self.assertEqual(len(invoice.items), 1)
        self.assertEqual(invoice.items[0].project_name, "服务费")
        self.assertEqual(invoice.items[0].tax_rate, Decimal("0.13"))
        self.assertTrue(all(not record.abnormal for record in invoice.validations))

    @patch("pdf_invoice_tool.convert_invoice_pdf")
    def test_batch_conversion_keeps_processing_after_one_invoice_fails(self, convert):
        convert.side_effect = [
            SimpleNamespace(output_file="first.xlsx", item_count=1, abnormal_count=0),
            ValueError("无法解析"),
            SimpleNamespace(output_file="third.xlsx", item_count=2, abnormal_count=1),
        ]

        results, failures = convert_invoice_pdfs(
            [self.root / "第一张.pdf", self.root / "损坏.pdf", self.root / "第三张.pdf"],
            self.root,
        )

        self.assertEqual(len(results), 2)
        self.assertEqual(Path(failures[0][0]).name, "损坏.pdf")
        self.assertEqual(convert.call_count, 3)

    def test_parser_keeps_inconsistent_values_for_validation(self):
        invoice = parse_invoice_blocks(
            invoice_blocks(["服务费", "1", "10000", "100.00", "13%", "13.00"])
        )

        item = invoice.items[0]
        self.assertEqual(item.quantity, Decimal("1"))
        self.assertEqual(item.unit_price, Decimal("10000"))
        self.assertEqual(item.amount, Decimal("100.00"))
        self.assertTrue(any(record.abnormal for record in invoice.validations))

    def test_party_fields_merge_continuation_lines_without_spaces(self):
        party = _extract_party(
            [
                "购买方",
                "名称: 购方公司",
                "地址、电话: 上海市浦东新区",
                "世纪大道100号 021-12345678",
                "开户行及账号: 测试银行",
                "123456789",
                "项目名称 数量 金额 税额",
            ],
            "购买方",
            ("项目名称", "销售方"),
        )

        self.assertEqual(
            party["地址电话"],
            "上海市浦东新区世纪大道100号 021-12345678",
        )
        self.assertEqual(party["开户行及账号"], "测试银行123456789")

    def test_positioned_parser_keeps_tax_exempt_row(self):
        blocks = [
            TextBlock("项目名称 数量 单价 金额 税率 税额", 1, 10, 100, 580),
            TextBlock("普通项目", 1, 10, 120, 100),
            TextBlock("1", 1, 270, 120, 280),
            TextBlock("100", 1, 330, 120, 350),
            TextBlock("100", 1, 400, 120, 430),
            TextBlock("13%", 1, 470, 120, 490),
            TextBlock("13", 1, 550, 120, 570),
            TextBlock("免税项目", 1, 10, 140, 100),
            TextBlock("1", 1, 270, 140, 280),
            TextBlock("50", 1, 330, 140, 350),
            TextBlock("50", 1, 400, 140, 430),
            TextBlock("免税", 1, 470, 140, 500),
            TextBlock("***", 1, 550, 140, 570),
            TextBlock("小计", 1, 50, 200, 100),
        ]

        items = _positioned_items(blocks)

        self.assertEqual([item.project_name for item in items], ["普通项目", "免税项目"])
        self.assertEqual(items[1].amount, Decimal("50"))
        self.assertIsNone(items[1].tax_amount)

    def test_multiple_invoices_are_rejected_instead_of_merged(self):
        blocks = []
        for page, number in ((1, "11111111"), (2, "22222222")):
            lines = [
                f"发票号码: {number}",
                "项目名称 数量 单价 金额 税率 税额",
                "服务费 1 100 100 13% 13",
            ]
            blocks.extend(
                TextBlock(text, page, top=index * 20)
                for index, text in enumerate(lines)
            )

        with self.assertRaisesRegex(ValueError, "检测到多张发票"):
            parse_invoice_blocks(blocks)

    def test_blank_page_is_not_treated_as_scanned_page(self):
        blank_page = SimpleNamespace(
            width=100,
            height=100,
            images=[],
            extract_words=lambda **kwargs: [],
        )
        scanned_page = SimpleNamespace(
            width=100,
            height=100,
            images=[{"width": 100, "height": 100}],
            extract_words=lambda **kwargs: [],
        )

        class FakeDocument:
            pages = [blank_page, scanned_page]

            def __enter__(self):
                return self

            def __exit__(self, *args):
                return False

        fake_pdfplumber = SimpleNamespace(open=lambda _: FakeDocument())
        with patch.dict("sys.modules", {"pdfplumber": fake_pdfplumber}):
            blocks, scanned_pages = _pdf_text_blocks("unused.pdf")

        self.assertEqual(blocks, [])
        self.assertEqual(scanned_pages, [2])

    def test_scanned_pdf_is_rejected_without_generating_output(self):
        source = self.root / "scan.pdf"
        source.write_bytes(b"%PDF-1.7")

        with patch(
            "pdf_invoice_tool._pdf_text_blocks",
            return_value=([], [1, 2]),
        ):
            with self.assertRaisesRegex(
                ScannedPdfUnsupportedError,
                "当前版本仅支持文本型 PDF",
            ):
                extract_invoice(source)

    def test_positioned_parser_handles_multipage_subtotal_and_side_by_side_parties(self):
        blocks = [
            TextBlock("名称：", 1, 30, 20, 60),
            TextBlock("购方公司", 1, 57, 20, 130),
            TextBlock("名称：", 1, 318, 20, 345),
            TextBlock("销方公司", 1, 342, 20, 420),
            TextBlock("项目名称 数量 单价 金额 税率 税额", 1, 10, 100, 580),
            TextBlock("服务费", 1, 10, 120, 260),
            TextBlock("1", 1, 270, 120, 280),
            TextBlock("100", 1, 330, 120, 350),
            TextBlock("100", 1, 400, 120, 430),
            TextBlock("13%", 1, 470, 120, 490),
            TextBlock("13", 1, 550, 120, 570),
            TextBlock("100", 1, 400, 200, 430),
            TextBlock("13", 1, 550, 200, 570),
            TextBlock("小计", 1, 50, 206, 100),
            TextBlock("项目名称 数量 单价 金额 税率 税额", 2, 10, 100, 580),
            TextBlock("200mm*200mm,70片/包", 2, 10, 112, 120),
            TextBlock("*调整项目*", 2, 10, 116, 80),
            TextBlock("-10", 2, 400, 120, 430),
            TextBlock("13%", 2, 470, 120, 490),
            TextBlock("-1.30", 2, 550, 120, 580),
            TextBlock("小计", 2, 50, 206, 100),
        ]

        invoice = parse_invoice_blocks(blocks)

        self.assertEqual(invoice.header["购买方名称"], "购方公司")
        self.assertEqual(invoice.header["销售方名称"], "销方公司")
        self.assertEqual(len(invoice.items), 2)
        self.assertEqual(
            invoice.items[0].project_name,
            "服务费200mm*200mm,70片/包",
        )
        self.assertEqual(invoice.items[0].specification, "")
        self.assertEqual(invoice.items[1].amount, Decimal("-10"))

    def test_positioned_party_parser_ignores_distant_download_counter(self):
        blocks = [
            TextBlock("名称：", 1, 321.5, 101.27, 348.5),
            TextBlock("销方公司", 1, 346.5, 100.58, 410.5),
            TextBlock("载", 1, 587.73, 99.4, 596.73),
            TextBlock("统一社会信用代码/纳税人识别号：", 1, 321.5, 129.27, 447.65),
            TextBlock("91441304MAC4RC1X94", 1, 443.5, 126.76, 573.1),
            TextBlock("：", 1, 587.73, 127.6, 596.73),
        ]

        _, seller = _parties_from_blocks(blocks)

        self.assertEqual(seller["名称"], "销方公司")
        self.assertEqual(seller["税号"], "91441304MAC4RC1X94")

    def test_validation_marks_inconsistent_fields_as_abnormal(self):
        invoice = InvoiceData()
        invoice.header.update(
            {
                "发票号码": "12345678",
                "合计金额（不含税）": Decimal("100"),
                "合计税额": Decimal("13"),
                "价税合计（小写）": Decimal("120"),
            }
        )
        invoice.items = [
            InvoiceItem(
                project_name="服务费",
                quantity=Decimal("2"),
                unit_price=Decimal("100"),
                amount=Decimal("100"),
                tax_rate=Decimal("0.13"),
                tax_amount=Decimal("10"),
            )
        ]

        records = validate_invoice(invoice)

        self.assertTrue(any(record.abnormal for record in records))
        abnormal_fields = "、".join(record.abnormal_fields for record in records if record.abnormal)
        self.assertIn("数量", abnormal_fields)
        self.assertIn("税额", abnormal_fields)
        self.assertIn("价税合计（小写）", abnormal_fields)

    def test_workbook_has_only_three_finance_ready_sheets(self):
        invoice = parse_invoice_blocks(invoice_blocks())
        output = self.root / "invoice.xlsx"

        write_invoice_workbook(invoice, output)

        workbook = load_workbook(output, data_only=False)
        try:
            self.assertEqual(workbook.sheetnames, ["发票头信息", "明细表", "校验结果"])
            self.assertEqual(
                tuple(cell.value for cell in workbook["发票头信息"][1]),
                HEADER_FIELDS,
            )
            self.assertEqual(
                tuple(cell.value for cell in workbook["明细表"][1]),
                ITEM_FIELDS,
            )
            self.assertIsInstance(workbook["明细表"]["F2"].value, (int, float))
            self.assertIsInstance(workbook["发票头信息"]["C2"].value, date)
            self.assertEqual(workbook["发票头信息"]["B2"].number_format, "@")
            self.assertEqual(workbook["校验结果"]["D2"].value, "正常")
            self.assertIsInstance(workbook["校验结果"]["H2"].value, (int, float))
            self.assertTrue(all(not sheet.tables for sheet in workbook.worksheets))
            self.assertEqual(workbook["明细表"].auto_filter.ref, "A1:H2")
            self.assertFalse(workbook["明细表"]["A2"].alignment.wrap_text)
            self.assertEqual(workbook["明细表"].column_dimensions["A"].width, 80)
            all_values = " ".join(
                str(cell.value or "")
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
            )
            self.assertNotIn("OCR原始文本", all_values)
        finally:
            workbook.close()
        with ZipFile(output) as archive:
            self.assertFalse(
                any(name.startswith("xl/tables/") for name in archive.namelist())
            )

    def test_invoice_ledger_workbook_opens_and_logs_key_fields(self):
        result = PdfInvoiceResult(
            output_file=str(self.root / "invoice.xlsx"),
            item_count=2,
            abnormal_count=1,
            source_file=str(self.root / "发票.pdf"),
            invoice_number="12345678",
            invoice_date="2026年6月23日",
            buyer_name="购方公司",
            seller_name="销方公司",
            seller_tax_id="91310000987654321X",
            amount=Decimal("100.00"),
            tax_amount=Decimal("13.00"),
            total_amount=Decimal("113.00"),
        )

        ledger = write_invoice_ledger(
            [result],
            [(self.root / "失败.pdf", "无法解析")],
            self.root,
        )

        workbook = load_workbook(ledger.output_file)
        try:
            sheet = workbook.active
            self.assertEqual(sheet.title, "发票台账")
            self.assertEqual(sheet["B2"].value, "12345678")
            self.assertEqual(sheet["F2"].value, 100)
            self.assertEqual(sheet["I2"].value, 1)
        finally:
            workbook.close()

        log_text = Path(ledger.log_file).read_text(encoding="utf-8")
        self.assertIn("匹配结果", log_text)
        self.assertIn("invoice_number=12345678", log_text)
        self.assertIn("seller_name=销方公司", log_text)
        self.assertIn("seller_tax_id=91310000987654321X", log_text)
        self.assertIn("失败 source_file=", log_text)
        self.assertIn("文件生成状态", log_text)

    def test_existing_output_is_preserved_without_confirmation(self):
        invoice = parse_invoice_blocks(invoice_blocks())
        output = self.root / "existing.xlsx"
        output.write_bytes(b"original")

        with self.assertRaisesRegex(FileExistsError, "输出文件已存在"):
            write_invoice_workbook(invoice, output)
        self.assertEqual(output.read_bytes(), b"original")

        write_invoice_workbook(invoice, output, overwrite=True)
        with ZipFile(output) as archive:
            self.assertIsNone(archive.testzip())

    def test_batch_output_uses_invoice_number_as_filename(self):
        source = self.root / "原发票.pdf"
        invoice = parse_invoice_blocks(invoice_blocks())

        with patch("pdf_invoice_tool.extract_invoice", return_value=invoice):
            first = convert_invoice_pdf(source, self.root)
            second = convert_invoice_pdf(source, self.root)

        self.assertEqual(Path(first.output_file).name, "12345678.xlsx")
        self.assertEqual(Path(second.output_file).name, "12345678_1.xlsx")

    def test_missing_amount_does_not_generate_unstructured_output(self):
        with self.assertRaisesRegex(ValueError, "包含“金额”的发票明细"):
            parse_invoice_blocks(invoice_blocks(["只有项目名称"]))


if __name__ == "__main__":
    unittest.main()
