import io
import json
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from unittest.mock import patch
from zipfile import ZipFile

from openpyxl import load_workbook

import document_router
from core.classifier import CONTRACT, INVOICE, TABLE, detect_doc_type
from core.document_router import process_document


class DocumentRouterTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.project_root = Path(__file__).resolve().parents[1]

    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)

    def tearDown(self):
        self.temporary_directory.cleanup()

    def test_root_module_remains_the_single_compatible_entry(self):
        self.assertIs(document_router.detect_doc_type, detect_doc_type)
        self.assertIs(document_router.process_document, process_document)

    def test_sample_documents_keep_the_same_types_and_outputs(self):
        expected = {
            "发票测试.pdf": (INVOICE, 0.95, ".xlsx"),
            "合同测试.pdf": (CONTRACT, 0.86, ".docx"),
            "表格测试.pdf": (TABLE, 0.83, ".xlsx"),
        }
        results = {}
        for filename, (doc_type, confidence, suffix) in expected.items():
            result = document_router.process_document(
                self.project_root / "test_files" / filename,
                self.root / "output",
                log_root=self.root / "logs",
            )
            self.assertEqual(result["status"], "success")
            self.assertEqual(result["doc_type"], doc_type)
            self.assertEqual(result["confidence"], confidence)
            self.assertEqual(Path(result["output_file"]).suffix, suffix)
            results[doc_type] = Path(result["output_file"])

        invoice = load_workbook(results[INVOICE], read_only=True)
        self.assertEqual(invoice.sheetnames, ["发票头信息", "明细表", "校验结果"])
        self.assertEqual(
            tuple(invoice["明细表"].iter_rows(min_row=2, max_row=2, values_only=True))[0],
            ("服务费", "标准", "项", 1, 100, 100, None, None),
        )
        invoice.close()

        table = load_workbook(results[TABLE], read_only=True)
        self.assertEqual(table.sheetnames, ["第1页_表1"])
        self.assertEqual(
            list(table.active.values),
            [
                ("项目名称", "数量", "单价", "金额"),
                ("办公用品", "2", "50.00", "100.00"),
                ("服务费用", "1", "300.00", "300.00"),
            ],
        )
        table.close()

        with ZipFile(results[CONTRACT]) as contract:
            document_xml = contract.read("word/document.xml").decode("utf-8")
        self.assertIn("采购合同", document_xml)
        self.assertNotIn("=== 第", document_xml)

    def test_unknown_text_output_remains_byte_compatible(self):
        source = self.root / "source.txt"
        source.write_text(
            "\n\n=== 第 1 页 ===\n\n这是一份普通说明文字\n",
            encoding="utf-8",
        )
        output = self.root / "unknown.txt"
        document_router._write_unknown_text(source, output)
        self.assertEqual(
            output.read_text(encoding="utf-8"),
            "文档类型：UNKNOWN\n识别结果：无法分类\n\n"
            "\n\n=== 第 1 页 ===\n\n这是一份普通说明文字\n",
        )

    def test_contract_page_break_and_table_empty_fields(self):
        contract_source = self.root / "contract.txt"
        contract_source.write_text(
            "=== 第 1 页 ===\n   第一页    内容   \n\n第一页第二段\n"
            "=== 第 2 页 ===\n第二页\n",
            encoding="utf-8",
        )
        contract_output = self.root / "contract.docx"
        document_router._write_contract_docx(
            contract_source,
            contract_output,
            self.root,
        )
        with ZipFile(contract_output) as contract:
            document_xml = contract.read("word/document.xml").decode("utf-8")
        self.assertEqual(document_xml.count('w:type="page"'), 1)
        self.assertNotIn("=== 第", document_xml)
        self.assertIn("第一页 内容", document_xml)
        self.assertNotIn("   第一页", document_xml)
        self.assertIn('w:line="60" w:lineRule="exact"', document_xml)

        self.assertEqual(
            document_router._valid_table(
                [["名称", "规格", "金额"], ["项目", None, "10"], ["其他", "标准"]]
            ),
            [["名称", "规格", "金额"], ["项目", None, "10"], ["其他", "标准", None]],
        )

    def test_contract_page_consensus_ignores_invoice_words_in_payment_terms(self):
        text = (
            "经销合作协议\n甲方：供应商\n乙方：经销商\f"
            "付款条件：乙方收到发票后付款，凭证应载明税号。\f"
            "协议条款\n甲方与乙方约定如下。\f"
            "结算说明：货款金额以对账结果为准，税额按适用政策执行。"
        )
        self.assertEqual(detect_doc_type(text), CONTRACT)

    def test_table_negative_numbers_and_failed_result_message(self):
        self.assertEqual(document_router._excel_safe("-123"), "-123")
        self.assertEqual(document_router._excel_safe("-1.25"), "-1.25")
        self.assertEqual(document_router._excel_safe("-1,234.50"), "-1,234.50")
        self.assertEqual(document_router._excel_safe("-cmd"), "'-cmd")
        self.assertEqual(document_router._excel_safe("=SUM(A1:A2)"), "'=SUM(A1:A2)")

        result = document_router.process_document(
            self.root / "missing.pdf",
            self.root / "output",
            log_root=self.root / "logs",
        )
        self.assertEqual(result["status"], "failed")
        self.assertIn("PDF 文件不存在", result["error_message"])

    def test_cli_keeps_json_stdout_and_reports_progress_on_stderr(self):
        expected = {
            "doc_type": CONTRACT,
            "confidence": 0.86,
            "output_file": "/tmp/合同.docx",
            "status": "success",
        }

        def fake_process(_pdf, _output, progress_callback, log_root):
            self.assertIsNone(log_root)
            progress_callback(1, 1, "正在读取第 1 页")
            return expected

        stdout = io.StringIO()
        stderr = io.StringIO()
        with patch("document_router.process_document", side_effect=fake_process):
            with redirect_stdout(stdout), redirect_stderr(stderr):
                exit_code = document_router.main(["合同.pdf"])

        self.assertEqual(exit_code, 0)
        self.assertEqual(json.loads(stdout.getvalue()), expected)
        self.assertIn("正在处理：合同.pdf", stderr.getvalue())
        self.assertIn("[1/1] 正在读取第 1 页", stderr.getvalue())
        self.assertIn("处理完成：/tmp/合同.docx", stderr.getvalue())


if __name__ == "__main__":
    unittest.main()
