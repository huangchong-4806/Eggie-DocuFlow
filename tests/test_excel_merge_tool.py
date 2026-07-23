import re
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from excel_merge_tool import (
    build_merged_workbook,
    discover_excel_files,
    get_file_info,
    get_workbook_metadata,
    split_workbook_by_rows,
)


class ExcelMergeToolTests(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_directory.name)

    def tearDown(self):
        self.temp_directory.cleanup()

    def create_workbook(self, filename, include_merge=False):
        path = self.root / filename
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.column_dimensions["A"].width = 22

        worksheet.append(["名称", "金额", "文本公式", "备注", None])
        worksheet.append([filename, 10, None, "合并内容", None])
        worksheet["B2"] = "=A2"
        worksheet["C2"] = "=保持文本"
        worksheet["C2"].data_type = "s"

        thin = Side(style="thin", color="000000")
        worksheet["A2"].font = Font(bold=True, color="FFFFFF")
        worksheet["A2"].fill = PatternFill("solid", fgColor="336699")
        worksheet["A2"].border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )
        worksheet["B2"].number_format = "#,##0.00"

        if include_merge:
            worksheet.merge_cells("D2:E2")

        workbook.save(path)
        workbook.close()
        return path

    def corrupt_dimension(self, filename):
        replacement = filename.with_suffix(".broken.xlsx")
        with ZipFile(filename) as source, ZipFile(
            replacement,
            "w",
            ZIP_DEFLATED,
        ) as target:
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    dimension_start = data.find(b"<dimension ")
                    dimension_end = data.find(b"/>", dimension_start)
                    data = (
                        data[:dimension_start]
                        + b'<dimension ref="A1"/>'
                        + data[dimension_end + 2 :]
                    )
                target.writestr(item, data)
        return replacement

    def remove_dimension(self, filename):
        replacement = filename.with_suffix(".no-dimension.xlsx")
        with ZipFile(filename) as source, ZipFile(
            replacement,
            "w",
            ZIP_DEFLATED,
        ) as target:
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    data, replacements = re.subn(
                        rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?dimension\b[^>]*/>",
                        b"",
                        data,
                        count=1,
                    )
                    self.assertEqual(replacements, 1)
                target.writestr(item, data)
        return replacement

    def test_metadata_reads_rows_from_broken_dimension(self):
        source = self.create_workbook("metadata.xlsx")
        broken = self.corrupt_dimension(source)

        metadata = get_workbook_metadata(broken)
        file_info = get_file_info(broken)

        self.assertEqual(metadata.row_count, 2)
        self.assertEqual(metadata.column_count, 4)
        self.assertEqual(file_info["rows"], 2)
        self.assertEqual(file_info["columns"], 4)
        self.assertEqual(file_info["merged_cells"], 0)
        self.assertIn((1, 1, 22.0), metadata.column_widths)

    def test_metadata_column_count_ignores_empty_width_only_columns(self):
        source = self.create_workbook("width-only-column.xlsx")
        workbook = load_workbook(source)
        try:
            worksheet = workbook.active
            worksheet.column_dimensions["Z"].width = 28
            workbook.save(source)
        finally:
            workbook.close()

        file_info = get_file_info(source)

        self.assertEqual(file_info["columns"], 4)

    def test_streaming_merge_preserves_values_formulas_and_styles(self):
        first = self.create_workbook("001.xlsx")
        second = self.create_workbook("002.xlsx")
        output = self.root / "stream-result.xlsx"

        build_merged_workbook([first, second], output, skip_rows=1)

        workbook = load_workbook(output, data_only=False)
        worksheet = workbook.active
        try:
            self.assertEqual(worksheet.max_row, 3)
            self.assertEqual(worksheet["A2"].value, "001.xlsx")
            self.assertEqual(worksheet["A3"].value, "002.xlsx")
            self.assertEqual(worksheet["B2"].value, "=A2")
            self.assertEqual(worksheet["B3"].value, "=A3")
            self.assertEqual(worksheet["C3"].value, "=保持文本")
            self.assertEqual(worksheet["C3"].data_type, "s")
            self.assertTrue(worksheet["A3"].font.bold)
            self.assertEqual(worksheet["A3"].fill.fgColor.rgb, "00336699")
            self.assertEqual(worksheet["B3"].number_format, "#,##0.00")
            self.assertEqual(worksheet.column_dimensions["A"].width, 22.0)
            self.assertEqual(len(worksheet.merged_cells.ranges), 0)
        finally:
            workbook.close()

    def test_streaming_merge_accepts_valid_file_without_dimension(self):
        first = self.remove_dimension(self.create_workbook("001.xlsx"))
        second = self.remove_dimension(self.create_workbook("002.xlsx"))
        output = self.root / "dimensionless-result.xlsx"

        build_merged_workbook([first, second], output, skip_rows=1)

        workbook = load_workbook(output, data_only=False)
        worksheet = workbook.active
        try:
            self.assertEqual(worksheet.max_row, 3)
            self.assertEqual(worksheet["A2"].value, "001.xlsx")
            self.assertEqual(worksheet["A3"].value, "002.xlsx")
        finally:
            workbook.close()

    def test_merged_cells_trigger_compatible_output(self):
        first = self.create_workbook("001.xlsx")
        second = self.create_workbook("002.xlsx", include_merge=True)
        output = self.root / "merged-result.xlsx"

        build_merged_workbook(
            [first, second],
            output,
            skip_rows=1,
            keep_merged_cells=True,
        )

        workbook = load_workbook(output, data_only=False)
        worksheet = workbook.active
        try:
            self.assertIn("D3:E3", worksheet.merged_cells)
            self.assertEqual(worksheet["D3"].value, "合并内容")
            self.assertEqual(worksheet["A3"].value, "002.xlsx")
        finally:
            workbook.close()

    def test_merged_cells_can_be_disabled_for_streaming_output(self):
        source = self.create_workbook("merged.xlsx", include_merge=True)
        output = self.root / "unmerged-result.xlsx"

        build_merged_workbook(
            [source],
            output,
            keep_merged_cells=False,
        )

        workbook = load_workbook(output)
        try:
            self.assertEqual(len(workbook.active.merged_cells.ranges), 0)
            self.assertEqual(workbook.active["D2"].value, "合并内容")
        finally:
            workbook.close()

    def test_merge_rejects_output_alias_of_source(self):
        source = self.create_workbook("source.xlsx")
        alias = self.root / "alias.xlsx"
        alias.symlink_to(source)

        with self.assertRaisesRegex(ValueError, "不能与待合并的源文件相同"):
            build_merged_workbook([source], alias)

        workbook = load_workbook(source)
        try:
            self.assertEqual(workbook.active.title, "Sheet")
        finally:
            workbook.close()

    def test_merge_failure_preserves_existing_output(self):
        source = self.create_workbook("atomic-source.xlsx")
        output = self.root / "existing-output.xlsx"
        output.write_bytes(b"existing result")

        def fail_after_writing(files, metadata, temporary_output, *args):
            Path(temporary_output).write_bytes(b"partial result")
            raise RuntimeError("save failed")

        with patch(
            "excel_merge_tool._build_streaming_workbook",
            side_effect=fail_after_writing,
        ):
            with self.assertRaisesRegex(RuntimeError, "save failed"):
                build_merged_workbook([source], output)

        self.assertEqual(output.read_bytes(), b"existing result")
        self.assertFalse(any(self.root.glob(".existing-output-*.xlsx")))

    def test_discover_excel_files_is_sorted_and_ignores_temporary_files(self):
        nested = self.root / "nested"
        nested.mkdir()
        self.create_workbook("B.xlsx")
        self.create_workbook("a.xlsx")
        temporary = self.root / "~$ignored.xlsx"
        temporary.write_bytes(b"temporary")
        hidden = nested / ".hidden.xlsx"
        hidden.write_bytes(b"hidden")

        discovered = discover_excel_files(self.root)

        self.assertEqual(
            [Path(filename).name for filename in discovered],
            ["a.xlsx", "B.xlsx"],
        )

    def test_split_workbook_repeats_headers_and_preserves_styles(self):
        source = self.root / "split-source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "明细"
        worksheet.column_dimensions["A"].width = 18
        worksheet.row_dimensions[1].height = 24
        worksheet.merge_cells("A1:C1")
        worksheet["A1"] = "订单明细"
        worksheet["A1"].font = Font(bold=True, color="FFFFFF")
        worksheet["A1"].fill = PatternFill("solid", fgColor="336699")
        thin = Side(style="thin", color="000000")
        worksheet["A1"].border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )
        worksheet.append(["订单号", "金额", "公式"])
        for index in range(1, 6):
            source_row = index + 2
            worksheet.append([f"订单{index}", index * 10, f"=B{source_row}*2"])
            worksheet[f"B{source_row}"].number_format = "#,##0.00"
        workbook.save(source)
        workbook.close()

        output_folder = self.root / "split-output"
        output_folder.mkdir()

        split_result = split_workbook_by_rows(
            source,
            output_folder,
            rows_per_file=2,
            header_rows=2,
        )
        output_files = split_result.output_files

        self.assertEqual(
            Path(split_result.output_folder),
            output_folder / "split-source_拆分结果",
        )
        self.assertEqual(split_result.total_rows, 7)
        self.assertEqual(split_result.header_rows, 2)
        self.assertEqual(split_result.data_rows, 5)
        self.assertEqual(split_result.file_count, 3)
        self.assertGreaterEqual(split_result.elapsed_seconds, 0)

        self.assertEqual(
            [Path(filename).name for filename in output_files],
            [
                "split-source_拆分001.xlsx",
                "split-source_拆分002.xlsx",
                "split-source_拆分003.xlsx",
            ],
        )
        self.assertEqual(
            {Path(filename).parent for filename in output_files},
            {Path(split_result.output_folder)},
        )
        self.assertFalse(any(output_folder.glob("*.xlsx")))

        expected_values = [
            ("订单1", "订单2", 4),
            ("订单3", "订单4", 4),
            ("订单5", None, 3),
        ]
        for filename, (first_order, second_order, max_row) in zip(
            output_files,
            expected_values,
        ):
            result = load_workbook(filename)
            try:
                result_sheet = result.active
                self.assertEqual(result_sheet.title, "明细")
                self.assertEqual(result_sheet.max_row, max_row)
                self.assertEqual(result_sheet["A1"].value, "订单明细")
                self.assertEqual(result_sheet["A2"].value, "订单号")
                self.assertEqual(result_sheet["A3"].value, first_order)
                if second_order is not None:
                    self.assertEqual(result_sheet["A4"].value, second_order)
                self.assertTrue(result_sheet["A1"].font.bold)
                self.assertEqual(result_sheet["A1"].fill.fgColor.rgb, "00336699")
                self.assertEqual(result_sheet["A1"].border.left.style, "thin")
                self.assertEqual(result_sheet.column_dimensions["A"].width, 18.0)
                self.assertEqual(result_sheet.row_dimensions[1].height, 24.0)
                self.assertEqual(result_sheet["B3"].number_format, "#,##0.00")
                self.assertEqual(result_sheet["C3"].value, "=B3*2")
                self.assertEqual(result_sheet["C3"].data_type, "f")
                self.assertIn("A1:C1", result_sheet.merged_cells)
            finally:
                result.close()

    def test_split_workbook_validates_user_inputs(self):
        source = self.create_workbook("split-validation.xlsx")
        output_folder = self.root / "split-validation-output"
        output_folder.mkdir()

        with self.assertRaisesRegex(ValueError, "每个文件的数据行数必须大于 0"):
            split_workbook_by_rows(source, output_folder, rows_per_file=0)

        with self.assertRaisesRegex(ValueError, "表头行数不能小于 0"):
            split_workbook_by_rows(source, output_folder, rows_per_file=1, header_rows=-1)

        with self.assertRaisesRegex(ValueError, "表头行数不能大于总行数"):
            split_workbook_by_rows(source, output_folder, rows_per_file=1, header_rows=3)

        with self.assertRaisesRegex(ValueError, "没有可拆分数据"):
            split_workbook_by_rows(source, output_folder, rows_per_file=1, header_rows=2)

        unsupported = self.root / "unsupported.xlsm"
        unsupported.write_bytes(b"not used")
        with self.assertRaisesRegex(ValueError, "只支持 .xlsx"):
            split_workbook_by_rows(unsupported, output_folder, rows_per_file=1)

    def test_split_workbook_uses_numbered_result_folder_when_needed(self):
        source = self.create_workbook("repeat-folder.xlsx")
        output_folder = self.root / "repeat-output"
        output_folder.mkdir()
        existing_folder = output_folder / "repeat-folder_拆分结果"
        existing_folder.mkdir()

        split_result = split_workbook_by_rows(
            source,
            output_folder,
            rows_per_file=1,
            header_rows=1,
        )

        self.assertEqual(
            Path(split_result.output_folder),
            output_folder / "repeat-folder_拆分结果_1",
        )
        self.assertTrue(existing_folder.is_dir())
        self.assertEqual(
            [Path(filename).parent for filename in split_result.output_files],
            [Path(split_result.output_folder)],
        )

    def test_split_rejects_merged_cells_across_part_boundary(self):
        source = self.root / "cross-boundary.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["表头"])
        worksheet["A2"] = "合并内容"
        worksheet.merge_cells("A2:A4")
        workbook.save(source)
        workbook.close()
        output_folder = self.root / "cross-boundary-output"
        output_folder.mkdir()

        with self.assertRaisesRegex(ValueError, "跨越拆分边界"):
            split_workbook_by_rows(
                source,
                output_folder,
                rows_per_file=2,
                header_rows=1,
            )

        self.assertEqual(list(output_folder.iterdir()), [])

    def test_split_failure_removes_partial_result_folder(self):
        source = self.create_workbook("partial-split.xlsx")
        output_folder = self.root / "partial-split-output"
        output_folder.mkdir()

        def fail_after_first_file(*args):
            raise RuntimeError("progress failed")

        with self.assertRaisesRegex(RuntimeError, "progress failed"):
            split_workbook_by_rows(
                source,
                output_folder,
                rows_per_file=1,
                header_rows=1,
                progress_callback=fail_after_first_file,
            )

        self.assertEqual(list(output_folder.iterdir()), [])


if __name__ == "__main__":
    unittest.main()
