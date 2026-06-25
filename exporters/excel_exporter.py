from pathlib import Path

from utils.file_helper import publish_output, temporary_output


def export_invoice(invoice, output_file):
    from pdf_invoice_tool import write_invoice_workbook

    return write_invoice_workbook(invoice, output_file)


def export_tables(tables, output_file):
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError("缺少 PDF 或 Excel 处理组件。") from error

    workbook = Workbook(write_only=True)
    table_count = 0
    temporary_file = None
    try:
        for page_number, page_table_number, table in tables:
            table_count += 1
            title = f"第{page_number}页_表{page_table_number}"[:31]
            sheet = workbook.create_sheet(title)
            width = len(table[0])
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = (
                f"A1:{get_column_letter(width)}{len(table)}"
            )
            for column, values in enumerate(zip(*table), 1):
                length = max(len(str(value or "")) for value in values)
                sheet.column_dimensions[get_column_letter(column)].width = min(
                    max(length + 2, 10), 50
                )
            for row_number, row in enumerate(table, 1):
                cells = []
                for value in row:
                    cell = WriteOnlyCell(sheet, value=value)
                    cell.alignment = Alignment(
                        horizontal="center" if row_number == 1 else "left",
                        vertical="center" if row_number == 1 else "top",
                        wrap_text=True,
                    )
                    if row_number == 1:
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.fill = PatternFill("solid", fgColor="17365D")
                    cells.append(cell)
                sheet.append(cells)

        if not table_count:
            raise ValueError("未提取到可用的行列结构。")
        temporary_file = temporary_output(output_file)
        workbook.save(temporary_file)
        return publish_output(temporary_file, output_file)
    finally:
        workbook.close()
        if temporary_file:
            Path(temporary_file).unlink(missing_ok=True)
